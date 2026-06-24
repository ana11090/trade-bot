"""
batch_compare_reports.py — read a folder of MT5 Strategy-Tester reports (.xlsx) and
line each one up against the matching Python backtest trades, producing one summary
table: per EA, MT5 entries vs Python entries, exact-bar matches, and one-bar shifts.

WHY: doing this by hand (open each report, count, eyeball against Python) doesn't
scale. This consumes the reports the batch runner produced and tells you which EAs
actually match Python.

Matching MT5 report -> Python trades:
  - MT5 report file name is the EA name (from the batch manifest's 'name').
  - Python trades for that rule come from a per-rule trades CSV/JSON you point at,
    keyed by the same name or rule_combo. Adjust _python_trades_for() to your export.

CHANGED: June 2026 — batch report comparison
Python 3.9+; needs openpyxl (already used elsewhere in the project).
"""

import os
import sys
import csv
import glob
import json
from datetime import datetime, timedelta

_HERE = os.path.dirname(os.path.abspath(__file__))
_ROOT = os.path.dirname(_HERE)


# CHANGED: June 2026 — pull Python trades from stored outputs/rules/*.json files
# WHY: trades already exist on disk from backtesting; no manual export needed.
#   Each rule_*_{combo}_{exit}_{hash}_{TF}.json has a trades list with full detail.
import re


def _py_rules_dir():
    """Return the project2_backtesting/outputs/rules directory."""
    return os.path.join(_ROOT, "project2_backtesting", "outputs", "rules")


def _norm(s):
    """Normalize string for matching: alphanumeric lowercase only."""
    return "".join(ch.lower() for ch in str(s) if ch.isalnum())


def _load_py_trades_for(combo, exit_name, entry_tf):
    """Find the stored rule JSON whose combo + exit + tf match.

    Returns: (trades_list, meta_dict) or (None, None)
    meta_dict includes: {"file": filename, "py_tf": stored_tf, "tf_match": bool}

    WHY: Must match combo + exit + TF. Comparing M5 EA vs H4 Python is meaningless.
    """
    d = _py_rules_dir()
    if not os.path.isdir(d):
        return None, None

    # Build safe combo prefix using the same sanitization as strategy_backtester's write path.
    # WHY: re.search(r"[0-9a-f]{8}") fails for combos like "#15_BUY_M5_4c_6179_ATR_Only_c2db"
    #      where underscores split the hex segments — no 8-char continuous match exists,
    #      so hexid=None and the filter falls through to a wrong rule's file.
    # CHANGED: June 2026 — safe-combo filename match replaces hex extraction
    _safe_combo = str(combo).lstrip('#')
    for _ch in (' ', '/', '\\', ':', '*', '?', '"', '<', '>', '|'):
        _safe_combo = _safe_combo.replace(_ch, '_')
    safe_combo = _safe_combo.lower()
    want_exit = _norm(exit_name)
    want_tf = _norm(entry_tf)

    best = None
    for f in glob.glob(os.path.join(d, "rule_*.json")):
        base = os.path.basename(f).lower()
        # Quick filter: filename must contain the full safe combo string
        if safe_combo not in base:
            continue

        try:
            with open(f, encoding="utf-8") as fh:
                rd = json.load(fh)
        except Exception:
            continue

        if not rd.get("trades"):
            continue

        r_exit = _norm(rd.get("exit_name") or rd.get("exit_strategy") or "")
        r_tf = _norm(rd.get("entry_tf") or "")

        # Score the match: exit match + tf match
        score = 0
        if want_exit and want_exit in r_exit:
            score += 2
        if want_tf and want_tf == r_tf:
            score += 2
        elif want_tf and want_tf in r_tf:
            score += 1

        if best is None or score > best[0]:
            best = (score, rd, f, r_tf)

    if not best:
        return None, None

    _, rd, f, r_tf = best
    meta = {
        "file": os.path.basename(f),
        "py_tf": rd.get("entry_tf"),
        "tf_match": (_norm(rd.get("entry_tf") or "") == want_tf),
        # run_max_spread_pips is embedded at write time (strategy_backtester).
        # 0/absent = backtest ran without spread filter (parity gap vs EA at 65 pips).
        "py_max_spread": rd.get("run_max_spread_pips", 0),
    }
    return rd["trades"], meta


def _parse_mt5_html_stats(path):
    """Parse summary stats from an MT5 HTML report (UTF-16). Returns dict or None."""
    # WHY: MT5 /config tester writes HTML, not xlsx. Stats live in <td> pairs.
    # CHANGED: June 2026 — HTML report parser replaces xlsx reader for summary stats
    # CHANGED: June 2026 — added data-config fields (symbol, period, bars, deposit, broker)
    import re
    for enc in ("utf-16", "utf-16-le", "utf-8"):
        try:
            with open(path, encoding=enc, errors="ignore") as f:
                html = f.read()
            break
        except Exception:
            html = ""
    if not html:
        return None

    def _grab(label):
        m = re.search(re.escape(label) + r'\s*:?\s*</td>\s*<td[^>]*>\s*([-\d., ]+)',
                      html, re.IGNORECASE)
        if not m:
            return None
        return m.group(1).replace(' ', '').replace(',', '')

    def _grab_text(label):
        """Tolerant grab for header fields — matches 'Label: value' or '<td>Label</td><td>value'."""
        # WHY: header fields aren't always strict <td> pairs; looser pattern for text values.
        m = re.search(re.escape(label) + r'\s*:?\s*(?:</td>\s*<td[^>]*>)?\s*([^<\n\r]+)',
                      html, re.IGNORECASE)
        return m.group(1).strip() if m else None

    return {
        "net_profit":    _grab("Total Net Profit") or _grab("Net Profit"),
        "profit_factor": _grab("Profit Factor"),
        "trades":        _grab("Total Trades"),
        # CHANGED: June 2026 — data-config fields (symbol, period, bars, deposit, broker)
        # WHY: confirms both sides tested same instrument/window/costs.
        "symbol":        _grab_text("Symbol"),
        "period":        _grab_text("Period"),
        "bars":          _grab_text("Bars"),
        "initial_deposit": _grab_text("Initial Deposit") or _grab_text("Initial deposit"),
        "broker":        _grab_text("Broker") or _grab_text("Company") or _grab_text("Server"),
    }


def _read_mt5_entries_html(path):
    """Return sorted list of MT5 entry datetimes from an HTML report (deals 'in')."""
    # WHY: MT5 /config writes HTML; deal rows look like <td>2026.01.05 08:00</td>...<td>in</td>
    # CHANGED: June 2026 — parse HTML trade table instead of xlsx
    import re
    for enc in ("utf-16", "utf-16-le", "utf-8"):
        try:
            with open(path, encoding=enc, errors="ignore") as f:
                html = f.read()
            break
        except Exception:
            html = ""
    if not html:
        return []
    # Each row: sequence of <td> cells; we want rows where one cell matches 'in' (entry direction)
    rows = re.findall(r'<tr[^>]*>(.*?)</tr>', html, re.DOTALL | re.IGNORECASE)
    out = []
    for row in rows:
        cells = re.findall(r'<td[^>]*>(.*?)</td>', row, re.DOTALL | re.IGNORECASE)
        cells = [re.sub(r'<[^>]+>', '', c).strip() for c in cells]
        if not cells:
            continue
        if any(c.lower() == 'in' for c in cells):
            # first cell that looks like a datetime
            for c in cells:
                try:
                    out.append(datetime.strptime(c[:16], '%Y.%m.%d %H:%M'))
                    break
                except Exception:
                    pass
    return sorted(out)


def _read_mt5_entries(path):
    """Dispatch to HTML or xlsx reader based on extension."""
    ext = os.path.splitext(path)[1].lower()
    if ext in ('.htm', '.html'):
        return _read_mt5_entries_html(path)
    # Legacy xlsx path (kept for back-compat)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        rows = list(wb.worksheets[0].iter_rows(values_only=True))
        hdr_i = None
        for i, row in enumerate(rows):
            if row and row[0] == 'Time' and 'Deal' in [str(c) for c in row]:
                hdr_i = i
                break
        if hdr_i is None:
            return []
        out = []
        for row in rows[hdr_i + 1:]:
            if not row or row[0] is None:
                continue
            if 'Balance:' in str(row[0]):
                break
            if len(row) > 4 and str(row[4]).lower() == 'in':
                try:
                    out.append(datetime.strptime(str(row[0])[:16], '%Y.%m.%d %H:%M'))
                except Exception:
                    pass
        return sorted(out)
    except Exception:
        return []


def _parse_mt5_html(path):
    """Return (trades, stats) from an MT5 HTML report. Full per-trade deals + summary stats.

    The detailed HTML report (635KB+) contains:
    - Summary stats table: Symbol, Period, Initial Deposit, Total Net Profit, Profit Factor, etc.
    - Deals table with columns: Time | Deal | Symbol | Type | Direction | Volume | Price | Order | Commission | Swap | Profit | Balance | Comment

    This parser extracts both stats and pairs in/out deals into complete trades.
    """
    import re

    # Read HTML with multiple encoding attempts
    html = ""
    for enc in ("utf-16", "utf-16-le", "utf-8"):
        try:
            with open(path, encoding=enc, errors="ignore") as f:
                html = f.read()
            if html:
                break
        except Exception:
            html = ""

    # Parse all table rows and cells first
    rows = re.findall(r'<tr[^>]*>(.*?)</tr>', html, re.DOTALL | re.I)

    def _cells(row):
        """Extract <td> cell contents from a table row, stripping HTML tags."""
        cs = re.findall(r'<td[^>]*>(.*?)</td>', row, re.DOTALL | re.I)
        return [re.sub(r'<[^>]+>', '', c).strip() for c in cs]

    def _all_text(row):
        """Extract both <td> and <th> cell text from a row (for stats scan)."""
        cs = re.findall(r'<t[dh][^>]*>(.*?)</t[dh]>', row, re.DOTALL | re.I)
        return [re.sub(r'<[^>]+>', '', c).strip() for c in cs]

    # CHANGED: June 2026 — per-row stats scan using _all_text (includes <th> cells)
    # WHY: the flat all_cells approach has two failure modes:
    #   1. Some MT5 HTML versions use <th> for stat labels — _cells(<td> only) misses them.
    #   2. Cross-row contamination: "Symbol" in the deals header row (13 cols) was picked
    #      up first and the "next cell" in the flat list was "Type", not "XAUUSD".
    # Per-row scan stays within the same <tr> so label+value must be adjacent in one row.
    def _stat(label):
        """Find label cell within the same <tr>, return next non-empty cell value."""
        lab = label.lower()
        for _row in rows:
            cs = _all_text(_row)
            for i, c in enumerate(cs):
                if c.strip().rstrip(":").lower() == lab:
                    for j in range(i + 1, min(i + 4, len(cs))):
                        v = cs[j].strip()
                        if v and v != ":":
                            # Strip thousand separators (space or non-breaking space)
                            return v.replace(" ", "").replace("\xa0", "")
        return None

    stats = {
        "symbol":          _stat("Symbol"),
        "period":          _stat("Period"),
        "bars":            _stat("Bars"),
        "initial_deposit": _stat("Initial Deposit"),
        "net_profit":      _stat("Total Net Profit"),
        "profit_factor":   _stat("Profit Factor"),
        "total_deals":     _stat("Total Deals"),
        "broker":          _stat("Broker") or _stat("Company") or _stat("Server"),
    }

    # Find the deals header row: contains 'Time' and 'Deal' and 'Direction'
    col = None
    start = 0
    for i, row in enumerate(rows):
        cs = _cells(row)
        if cs and "Time" in cs and "Deal" in cs and ("Direction" in cs or "Profit" in cs):
            # Map column names to indices
            col = {name: k for k, name in enumerate(cs)}
            start = i + 1
            break

    trades = []
    if col is not None:
        def g(cs, name):
            """Get cell value by column name."""
            k = col.get(name)
            return cs[k] if (k is not None and k < len(cs)) else None

        def _num(x):
            """Parse numeric value, handling spaces and non-breaking spaces."""
            try:
                return float(str(x).replace(" ", "").replace("\xa0", ""))
            except Exception:
                return None

        # Pair in/out deals
        pend = None
        for row in rows[start:]:
            cs = _cells(row)
            if not cs:
                continue

            dirn = (g(cs, "Direction") or "").lower()
            typ = (g(cs, "Type") or "").lower()

            # Skip balance adjustments
            if typ == "balance":
                continue

            if dirn == "in":
                # Entry deal - save for pairing
                pend = cs
            elif dirn == "out" and pend is not None:
                # Exit deal - pair with pending entry
                trades.append({
                    "entry_time":  g(pend, "Time"),
                    "exit_time":   g(cs, "Time"),
                    "entry_price": _num(g(pend, "Price")),
                    "exit_price":  _num(g(cs, "Price")),
                    "direction":   g(pend, "Type"),
                    "volume":      g(pend, "Volume"),
                    "commission":  (_num(g(pend, "Commission")) or 0) + (_num(g(cs, "Commission")) or 0),
                    "swap":        _num(g(cs, "Swap")) or 0,
                    "profit":      _num(g(cs, "Profit")) or 0,
                    "comment":     g(pend, "Comment") or "",  # capture for future indicator logging
                })
                pend = None

    return trades, stats


# CHANGED: June 2026 — full MT5 xlsx parser (trades + stats)
# WHY: MT5 tester writes ReportTester xlsx with full deal pairs (in/out) and summary stats.
#   The HTML report lacks per-trade details; xlsx has entry/exit times, prices, profit, commission.
def _parse_mt5_xlsx(path):
    """Return (trades, stats) from an MT5 ReportTester xlsx.

    trades: list of dicts with entry_time, exit_time, entry_price, exit_price,
            direction, volume, profit, commission, swap.
    stats:  dict with symbol, period, initial_deposit, net_profit, profit_factor, total_deals.
    """
    import openpyxl
    import warnings
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # CHANGED: June 2026 — search all worksheets for stats (labels may be on different sheet)
    # WHY: some MT5 builds put summary on a separate sheet; iterate all sheets for robustness.
    rows = []
    for ws in wb.worksheets:
        rows += list(ws.iter_rows(values_only=True))

    def _cellval(label):
        """Find a label cell and return the next non-empty cell's value."""
        for r in rows:
            cells = list(r)
            for i, c in enumerate(cells):
                if c is not None and str(c).strip().rstrip(":").lower() == label.lower():
                    # Value is the next non-None cell
                    for j in range(i + 1, len(cells)):
                        if cells[j] is not None and str(cells[j]).strip():
                            return str(cells[j]).strip()
        return None

    stats = {
        "symbol":          _cellval("Symbol"),
        "period":          _cellval("Period"),
        "initial_deposit": _cellval("Initial Deposit"),
        "net_profit":      _cellval("Total Net Profit"),
        "profit_factor":   _cellval("Profit Factor"),
        "total_deals":     _cellval("Total Deals"),
    }

    # CHANGED: June 2026 — find the deals HEADER row like the working entries reader
    # WHY: looking for 'Deals' section label didn't match some MT5 layouts; the proven reader
    #   finds the header by 'Time' in first cell + 'Deal' somewhere in the row. This IS the
    #   column header row, so data starts at hdr+1.
    # CHANGED: June 2026 — accept 'Time' anywhere in row (not just first cell) for robustness
    # WHY: some MT5 layouts may have a different column order or leading columns.
    hdr_idx = None
    for i, r in enumerate(rows):
        if not r:
            continue
        cells = [str(c).strip() for c in r if c is not None]
        # Match if row contains both 'Time' AND 'Deal' (column headers)
        if "Time" in cells and "Deal" in cells:
            hdr_idx = i
            break

    trades = []
    if hdr_idx is None:
        return trades, stats  # No deals header found, return empty trades

    header = [str(c).strip() if c is not None else "" for c in rows[hdr_idx]]
    col = {name: k for k, name in enumerate(header)}

    def g(row, name):
        """Get column value by name."""
        k = col.get(name)
        return row[k] if (k is not None and k < len(row)) else None

    pending_in = None
    for r in rows[hdr_idx + 1:]:
        if not r or r[0] is None:
            continue
        # Stop at Balance: summary row (marks end of deals in some layouts)
        if "Balance:" in str(r[0]):
            break
        typ = str(g(r, "Type") or "").lower()
        direction = str(g(r, "Direction") or "").lower()
        # Skip balance rows
        if typ == "balance":
            continue
        # Pair in/out deals into trades
        if direction == "in":
            pending_in = r
        elif direction == "out" and pending_in is not None:
            trades.append({
                "entry_time":  g(pending_in, "Time"),
                "exit_time":   g(r, "Time"),
                "entry_price": g(pending_in, "Price"),
                "exit_price":  g(r, "Price"),
                "direction":   str(g(pending_in, "Type") or ""),
                "volume":      g(pending_in, "Volume"),
                "commission":  (float(g(pending_in, "Commission") or 0) +
                               float(g(r, "Commission") or 0)),
                "swap":        float(g(r, "Swap") or 0),
                "profit":      float(g(r, "Profit") or 0),
            })
            pending_in = None

    return trades, stats


# CHANGED: June 2026 — find MT5 report with xlsx preference
# WHY: xlsx has full deal data; htm is summary-only fallback.
# CHANGED: June 2026 — sort by newest (mtime) to get the right report when multiple exist.
def _find_report(reports_dir, ea_name):
    """Find the MT5 report for an EA, preferring xlsx over htm/html, newest first."""
    import glob as _glob
    # Try exact name match first (EA name + extension), newest first
    for pat in (ea_name + ".xlsx",
                "*" + ea_name + "*.xlsx",      # partial match (e.g. ReportTester-<n>.xlsx)
                "ReportTester*.xlsx",           # MT5 default name if EA-named path wasn't honored
                ea_name + ".htm",
                ea_name + ".html"):
        hits = sorted(_glob.glob(os.path.join(reports_dir, pat)),
                      key=lambda p: os.path.getmtime(p), reverse=True)
        if hits:
            return hits[0]  # Return newest match
    return None


def _read_python_entries(csv_path):
    """Return sorted list of Python entry datetimes (minute-truncated)."""
    out = []
    with open(csv_path, encoding='utf-8') as f:
        for r in csv.DictReader(f):
            try:
                out.append(datetime.strptime(r['entry_time'], '%Y-%m-%d %H:%M:%S')
                           .replace(second=0))
            except Exception:
                pass
    return sorted(out)


# CHANGED: June 2026 — extract entry times from stored trade list (not CSV)
# WHY: _load_py_trades_for returns a list of trade dicts; we need datetimes for comparison.
def _extract_py_entry_times(trades):
    """Return sorted list of Python entry datetimes from a trade list."""
    out = []
    if not trades:
        return out
    for t in trades:
        if not isinstance(t, dict):
            continue
        et = t.get('entry_time')
        if not et:
            continue
        try:
            # handle both string and datetime formats
            if isinstance(et, str):
                # WHY: Vectorized FixedSLTP stores entry_time as numpy.datetime64,
                #      which serializes to ISO format "2026-01-14T16:00:00" (T separator).
                #      The iterative path stores "2026-01-14 16:00:00" (space separator).
                #      fromisoformat handles both; strptime only handles one.
                # CHANGED: June 2026 — handle ISO T-separator from vectorized path
                dt = datetime.fromisoformat(et).replace(second=0)
            elif isinstance(et, datetime):
                dt = et.replace(second=0)
            else:
                continue
            out.append(dt)
        except Exception:
            pass
    return sorted(out)


def _clean_combo(combo):
    # WHY: mirror view_results.py:992 exactly so names line up.
    # CHANGED: June 2026
    return str(combo).replace(' ', '_').replace('/', '_')[:30]


def _clean_exit(ex):
    # WHY: mirror view_results.py:993 exactly.
    # CHANGED: June 2026
    return str(ex).replace(' ', '_').replace('/', '_')[:20]


def _python_trades_for(name, python_dir, rec=None):
    """Locate the Python trades CSV for an EA.

    Preferred path (rec given from manifest): reconstruct the backtester's export
    name  trades_{clean_combo}_{clean_exit}{_tf}_{stamp}.csv  and glob the stamp,
    returning the NEWEST match. Falls back to the old name-based candidates.
    """
    # WHY: manifest-driven match against the real exporter naming scheme.
    # CHANGED: June 2026 — view_results exports trades_{combo}_{exit}{_tf}_{stamp}.csv
    if rec:
        combo = rec.get('rule_combo') or name
        ex = rec.get('exit_name') or ''
        tf = rec.get('entry_tf') or ''
        tf_tag = ('_' + tf) if tf else ''
        pat = 'trades_%s_%s%s_*.csv' % (_clean_combo(combo), _clean_exit(ex), tf_tag)
        hits = glob.glob(os.path.join(python_dir, pat))
        if hits:
            # newest run wins (the stamp guard means many runs may coexist)
            hits.sort(key=os.path.getmtime, reverse=True)
            return hits[0]

    # CHANGED: June 2026 — tolerant match: extract combo (8-hex) + exit keyword from EA name
    # WHY: EA names now carry exit+tf suffixes (e.g. BUY_M5_4c_0608_198587b7_M5_ATR_Only),
    #   while CSVs are trades_{combo}_{exit}{tf}_{stamp}.csv. Match by combo + exit token.
    # CHANGED: June 2026 — guard against empty python_dir (when no CSVs exported yet)
    if not python_dir:
        return None
    toks = name.replace("BUY_", "").replace("SELL_", "").split("_")
    # find the 8-hex combo token (e.g. 198587b7)
    combo = next((t for t in toks if len(t) == 8 and all(c in "0123456789abcdef" for c in t.lower())), None)
    cand = glob.glob(os.path.join(python_dir, "trades_*.csv"))
    if combo:
        cand = [c for c in cand if combo in os.path.basename(c)]
    # further narrow by an exit keyword if present in the EA name
    for kw in ("ATR_Only", "ATR_Fixed", "Fixed_SL", "Time_Based", "Indicator", "Hybrid",
               "Trailing", "PSAR", "Breakeven"):
        if kw.lower() in name.lower():
            hit = [c for c in cand if kw.lower() in os.path.basename(c).lower()]
            if hit:
                cand = hit
                break
    # newest match wins
    if cand:
        cand.sort(key=lambda p: os.path.getmtime(p), reverse=True)
        return cand[0]

    # Fallback: legacy/simple layouts.
    cands = [
        os.path.join(python_dir, name + '.csv'),
        os.path.join(python_dir, name + '_trades.csv'),
        os.path.join(python_dir, 'trades_' + name + '.csv'),
    ]
    for c in cands:
        if os.path.exists(c):
            return c
    shared = os.path.join(python_dir, 'trades.csv')
    return shared if os.path.exists(shared) else None


_TF_MINUTES = {"D1": 1440, "H4": 240, "H1": 60, "M30": 30, "M15": 15, "M5": 5, "W1": 10080}

def _tf_bar_minutes(tf):
    """Return the bar duration in minutes for an entry TF string (e.g. 'H4' → 240)."""
    return _TF_MINUTES.get((tf or "").upper(), 5)


def _compare(mt5, py, bar_minutes=5):
    """Return dict of comparison metrics between two sorted datetime lists."""
    mset, pset = set(mt5), set(py)
    if not mt5:
        return dict(mt5=0, py=len(py), py_in_range=0, exact=0,
                    one_bar_early=0, one_bar_late=0, py_after_end=0)
    cutoff = max(mt5)
    py_in = set(p for p in pset if p <= cutoff)
    delta = timedelta(minutes=bar_minutes)
    early = sum(1 for p in pset if (p + delta) in mset)   # python earlier than MT5
    late = sum(1 for p in pset if (p - delta) in mset)    # python later than MT5
    # WHY (June 2026): Session gap entries — Python enters at 00:00 (bar open),
    #      MT5 enters at 01:05 (first tick after reopen). These are 65 min apart,
    #      not matchable by exact or 1-bar offset. Count a Python midnight entry
    #      as "exact" if an MT5 entry exists within 120 min after it.
    # CHANGED: June 2026 — session gap tolerance for midnight entries
    exact_set = mset & pset
    for p in pset - exact_set:
        if p.hour == 0:
            for offset_min in range(5, 121, 5):
                if (p + timedelta(minutes=offset_min)) in mset:
                    exact_set.add(p)
                    break
    return dict(
        mt5=len(mset), py=len(pset), py_in_range=len(py_in),
        exact=len(exact_set),
        # WHY (June 2026 — matched P&L): Return the set of PY entry times that
        #      matched MT5 entries, so the caller can filter PY trades and compute
        #      P&L only for the trades both sides agree on.
        # CHANGED: June 2026 — return exact_set for matched-only P&L
        _exact_set=exact_set,
        one_bar_early=early, one_bar_late=late,
        py_after_end=len(pset - py_in),
    )


def compare_folder(reports_dir, python_dir, manifest_path=None, bar_minutes=5):
    """Compare every MT5 .xlsx report in reports_dir to its Python trades.

    Prints a table and returns the list of row dicts.
    """
    # WHY: keep full manifest record per name so we can rebuild the export filename,
    #      not just a display label.
    # CHANGED: June 2026
    label = {}
    rec_by_name = {}
    if manifest_path and os.path.exists(manifest_path):
        try:
            for rec in json.load(open(manifest_path, encoding='utf-8')):
                rec_by_name[rec['name']] = rec
                label[rec['name']] = rec.get('rule_combo', rec['name'])
        except Exception:
            pass

    # CHANGED: June 2026 — MT5 /config writes HTML reports; accept .htm and .html (+ legacy .xlsx)
    reports = sorted(glob.glob(os.path.join(reports_dir, '*.htm')) +
                     glob.glob(os.path.join(reports_dir, '*.html')) +
                     glob.glob(os.path.join(reports_dir, '*.xlsx')))
    if not reports:
        print('[CMP] no reports (.htm/.html/.xlsx) in %s' % reports_dir)
        return ["No MT5 reports found — check the tester actually wrote them."]

    rows = []
    print('\n%-32s %5s %5s %8s %6s %7s %6s %6s  %s' %
          ('EA', 'MT5', 'PY', 'net_pip', 'exact', 'early', 'late', 'after', 'note'))
    print('-' * 100)
    for rp in reports:
        name = os.path.splitext(os.path.basename(rp))[0]

        # CHANGED: June 2026 — use full parsers for stats (populate net_profit, profit_factor)
        # WHY: _parse_mt5_html and _parse_mt5_xlsx return complete stats; _parse_mt5_html_stats
        #   was a partial fallback. Full parsers extract summary stats from detailed reports.
        stats = {}
        try:
            if rp.lower().endswith('.xlsx'):
                _, stats = _parse_mt5_xlsx(rp)
            elif rp.lower().endswith(('.htm', '.html')):
                _, stats = _parse_mt5_html(rp)
        except Exception:
            stats = {}

        mt5 = _read_mt5_entries(rp)

        # CHANGED: June 2026 — load Python trades from stored outputs/rules/*.json
        # WHY: trades already exist on disk; no manual export/CSV needed. Fully automatic.
        #   Must match combo + exit + TF (comparing M5 EA vs H4 Python is meaningless).
        rec = rec_by_name.get(name, {})
        combo = str(rec.get('rule_combo') or '')
        exit_name = rec.get('exit_name') or ''
        tf = rec.get('entry_tf') or ''
        # WHY: one-bar offset for H4 rules is 4 h (240 min), not 5 min.
        #   Hardcoding bar_minutes=5 made one_bar_early/late blind to H4/H1 shifts.
        # CHANGED: 2026-06-17 — per-EA bar minutes derived from entry_tf
        per_bm = _tf_bar_minutes(tf)
        py_trades, meta = _load_py_trades_for(combo, exit_name, tf) if combo else (None, None)

        if not py_trades:
            # No stored trades for this rule (never backtested, or different combo)
            print('%-32s %5d   no stored Python run for this rule (combo=%s)' %
                  (name[:32], len(mt5), combo[:8] or '?'))
            rows.append(dict(name=name, mt5=len(mt5), error='no_python',
                            note='no stored Python run'))
            continue

        # Check TF match — if stored Python run is different TF than EA, comparison is meaningless
        if meta and not meta.get('tf_match'):
            py_tf = meta.get('py_tf') or '?'
            note = f"PY run TF={py_tf} ≠ EA TF={tf} (counts not comparable)"
            print('%-32s %5d %5d %8s  %s' %
                  (name[:32], len(mt5), len(py_trades),
                   (stats.get('net_profit') or '?')[:8], note))
            rows.append(dict(name=name, mt5=len(mt5), py=len(py_trades),
                            mt5_net_profit=stats.get('net_profit', ''),
                            mt5_profit_factor=stats.get('profit_factor', ''),
                            note=note))
            continue

        # TF matches → do real entry-time comparison
        py = _extract_py_entry_times(py_trades)
        m = _compare(mt5, py, bar_minutes=per_bm)
        m['name'] = name
        m['mt5_net_profit'] = stats.get('net_profit', '')
        m['mt5_profit_factor'] = stats.get('profit_factor', '')
        # WHY: Show Python profit alongside MT5 for at-a-glance comparison.
        # CHANGED: June 2026 — enriched comparison_summary for parity dashboard
        # WHY (June 2026 — in-range P&L): Total PY profit includes trades after
        #      MT5's window end, inflating the gap. Add in-range columns that only
        #      sum P&L from trades within the MT5 window for fair comparison.
        # CHANGED: June 2026 — in-range P&L columns + renamed total columns
        _py_npips = sum(float(t.get('net_pips', 0) or 0) for t in py_trades)
        _py_nprof = sum(float(t.get('net_profit', 0) or 0) for t in py_trades)
        _py_wins = sum(1 for t in py_trades if float(t.get('net_pips', 0) or 0) > 0)
        m['py_pips_total'] = round(_py_npips, 1)
        m['py_profit_total'] = round(_py_nprof, 2)
        m['py_win_rate'] = round(100.0 * _py_wins / len(py_trades), 1) if py_trades else 0
        # In-range P&L: only PY trades with entry_time <= last MT5 entry
        _cutoff = max(mt5) if mt5 else None
        if _cutoff:
            _ir_trades = []
            for _t in py_trades:
                try:
                    _et = _t.get('entry_time')
                    if isinstance(_et, str):
                        _et_dt = datetime.fromisoformat(_et)
                    elif isinstance(_et, datetime):
                        _et_dt = _et
                    else:
                        continue
                    if _et_dt.replace(second=0) <= _cutoff:
                        _ir_trades.append(_t)
                except Exception:
                    pass
            _ir_pips = sum(float(t.get('net_pips', 0) or 0) for t in _ir_trades)
            _ir_prof = sum(float(t.get('net_profit', 0) or 0) for t in _ir_trades)
            m['py_pips_inrange'] = round(_ir_pips, 1)
            m['py_profit_inrange'] = round(_ir_prof, 2)
        else:
            m['py_pips_inrange'] = m['py_pips_total']
            m['py_profit_inrange'] = m['py_profit_total']
        # WHY (June 2026 — matched P&L): Sum P&L only from PY trades whose entry
        #      time is in the exact-match set. This gives a true apples-to-apples
        #      comparison: same trades, same count, only per-trade pricing differs.
        # CHANGED: June 2026 — matched-only P&L column
        _exact_set = m.pop('_exact_set', set())
        if _exact_set:
            _matched_trades = []
            for _t in py_trades:
                try:
                    _et = _t.get('entry_time')
                    if isinstance(_et, str):
                        _et_dt = datetime.fromisoformat(_et).replace(second=0)
                    elif isinstance(_et, datetime):
                        _et_dt = _et.replace(second=0)
                    else:
                        continue
                    if _et_dt in _exact_set:
                        _matched_trades.append(_t)
                except Exception:
                    pass
            _m_pips = sum(float(t.get('net_pips', 0) or 0) for t in _matched_trades)
            _m_prof = sum(float(t.get('net_profit', 0) or 0) for t in _matched_trades)
            m['py_profit_matched'] = round(_m_prof, 2)
            m['py_pips_matched'] = round(_m_pips, 1)
            m['matched_trades'] = len(_matched_trades)
        else:
            m['py_profit_matched'] = m['py_profit_inrange']
            m['py_pips_matched'] = m['py_pips_inrange']
            m['matched_trades'] = m.get('exact', 0)
        m['mt5_only'] = m['mt5'] - m['py_in_range']
        m['py_only'] = m['py'] - m['py_in_range']
        m['parity_pct'] = round(100.0 * m['exact'] / m['mt5'], 0) if m['mt5'] else 0
        m['note'] = f"matched (file: {meta.get('file', '?')[:20]})"
        # Warn when Python backtest had no spread filter but EA uses 65 pips.
        # WHY: EA blocks entries when spread > 65 pips; Python without the filter
        #      keeps those bars — producing extra Python trades not in EA output.
        #      Fix: re-run backtest with "Use Config" ON + prop firm (leveraged).
        # CHANGED: June 2026 — spread filter mismatch warning
        _py_spread = (meta or {}).get('py_max_spread', 0)
        if not _py_spread:
            m['note'] += ' [SPREAD:PY=0 vs EA=65pip]'
            print(f'  [SPREAD WARN] {name[:40]}: Python run had max_spread_pips=0 '
                  f'(filter OFF). EA uses 65 pip gate. Re-run backtest with '
                  f'"Use Config" ON + leveraged firm to get parity.')
        rows.append(m)
        _gap_matched = ''
        try:
            _gap_matched = round(float(stats.get('net_profit', 0)) - m.get('py_profit_matched', 0), 2)
        except Exception:
            pass
        print('%-32s MT5=$%-8s PY_matched=$%-8s  gap=$%-8s  %d/%d exact (%d%%)' %
              (name[:32],
               (stats.get('net_profit') or '?')[:8],
               str(m.get('py_profit_matched', '?'))[:8],
               str(_gap_matched)[:8],
               m['exact'], m['mt5'], m['parity_pct']))

    # write CSV summary
    out_csv = os.path.join(reports_dir, 'comparison_summary.csv')
    keys = ['name', 'mt5_net_profit', 'py_profit_matched', 'py_profit_inrange',
            'py_profit_total', 'mt5_profit_factor',
            'mt5', 'py', 'py_in_range', 'matched_trades', 'exact', 'parity_pct',
            'mt5_only', 'py_only',
            'py_pips_matched', 'py_pips_inrange', 'py_pips_total', 'py_win_rate',
            'one_bar_early', 'one_bar_late', 'py_after_end', 'note']
    with open(out_csv, 'w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=keys, extrasaction='ignore')
        w.writeheader()
        for r in rows:
            w.writerow(r)
    print('\n[CMP] summary -> %s' % out_csv)
    print('Columns: exact=same-bar entries; early/late=Python off by one bar;')
    print('         after=Python trades past MT5 end (expected if Python ran longer).')
    print('         note=TF mismatch warning or match confirmation.')
    # NOTE: the eval-windows report is now produced by
    #       generate_eval_windows_report(), called from batch_eas_panel after the
    #       parity report (where the correct rules_dir is available). The older
    #       generate_eval_report() hook here was removed to avoid running the
    #       sliding-window simulation twice per batch.
    return rows


def compare_reports(reports_dir, python_dir, manifest_path=''):
    # WHY: panel wants printable lines, not dicts; capture compare_folder's stdout.
    # CHANGED: June 2026
    import io
    import contextlib
    buf = io.StringIO()
    with contextlib.redirect_stdout(buf):
        compare_folder(reports_dir, python_dir,
                       manifest_path=manifest_path or None)
    return buf.getvalue().splitlines()


# ── Eval Windows Report ──────────────────────────────────────────────────────
# WHY (June 2026): After parity comparison, auto-run sliding-window eval
#      simulation for each rule and produce an xlsx showing pass/fail per
#      window. Answers "would this rule pass the Get Leveraged challenge?"
# CHANGED: June 2026 — eval_windows_report feature

def generate_eval_windows_report(
    rules_dir,
    reports_dir,
    manifest_path=None,
    firm_id='leveraged',
    challenge_id='leveraged_standard',
    account_size=10000,
    risk_pct=1.0,
    sl_pips=150.0,
    pip_value_per_lot=1.0,
):
    """Generate eval_windows_report.xlsx from stored Python trade JSONs.

    Parameters
    ----------
    rules_dir     : path to outputs/rules/ with rule_*.json files
    reports_dir   : path to batch/reports/ (output goes in reports_dir/eval_windows/)
    manifest_path : optional batch_manifest.json for combo labels
    firm_id       : prop firm id (default 'leveraged')
    challenge_id  : challenge id (default 'leveraged_standard')
    account_size  : eval account size (default 10000)
    risk_pct      : risk per trade % (default 1.0)
    sl_pips       : SL distance in pips (default 150.0)
    pip_value_per_lot : $ per pip per lot (default 1.0 for XAUUSD)
    """
    import json
    import glob
    import os
    import pandas as pd
    from datetime import datetime

    try:
        from shared.prop_firm_simulator import simulate_challenge
        from project2_backtesting.strategy_validator import _trades_to_df
    except ImportError as e:
        print(f"[EVAL] Cannot import simulator: {e}")
        return None

    # Try openpyxl — if not available, skip xlsx generation
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("[EVAL] openpyxl not available — skipping eval_windows_report")
        return None

    # Load manifest for labels
    manifest = {}
    if manifest_path and os.path.exists(manifest_path):
        try:
            for rec in json.load(open(manifest_path, encoding='utf-8')):
                manifest[rec.get('name', '')] = rec
        except Exception:
            pass

    # Find all rule JSON files
    if not rules_dir or not os.path.isdir(rules_dir):
        print(f"[EVAL] Rules dir not found: {rules_dir}")
        return None

    rule_files = sorted(glob.glob(os.path.join(rules_dir, 'rule_*.json')))
    if not rule_files:
        print(f"[EVAL] No rule_*.json files in {rules_dir}")
        return None

    print(f"\n[EVAL] Running sliding-window eval for {len(rule_files)} rules "
          f"(firm={firm_id}, challenge={challenge_id}, acct=${account_size:,})")

    summary_rows = []
    window_rows = []

    for rf in rule_files:
        fname = os.path.basename(rf)
        try:
            with open(rf, encoding='utf-8') as f:
                rd = json.load(f)
        except Exception as e:
            print(f"  [EVAL] SKIP {fname}: {e}")
            continue

        trades = rd.get('trades') or rd.get('py_trades') or []
        if not trades or len(trades) < 3:
            print(f"  [EVAL] SKIP {fname}: {len(trades)} trades (need >=3)")
            continue

        rule_tf = rd.get('entry_tf', '')
        rule_combo = rd.get('rule_combo', fname)
        exit_config = rd.get('exit_name', '')

        # Convert trades to DataFrame for simulator
        df = _trades_to_df(trades, risk_pct, sl_pips, pip_value_per_lot, account_size)
        if df.empty:
            continue

        # Run simulation
        try:
            sim = simulate_challenge(
                trades_df=df,
                firm_id=firm_id,
                challenge_id=challenge_id,
                account_size=account_size,
                mode='sliding_window',
                simulate_funded=False,
                risk_per_trade_pct=risk_pct,
                default_sl_pips=sl_pips,
                pip_value_per_lot=pip_value_per_lot,
                symbol='XAUUSD',
            )
        except Exception as e:
            print(f"  [EVAL] ERROR {fname}: {e}")
            continue

        if sim is None:
            print(f"  [EVAL] SKIP {fname}: simulator returned None")
            continue

        # Compute max consecutive fails
        individual = getattr(sim, 'individual_results', []) or []
        max_consec = 0
        cur_consec = 0
        for r in individual:
            if r.eval_outcome in ('FAIL_DD', 'FAIL_DAILY_DD'):
                cur_consec += 1
                max_consec = max(max_consec, cur_consec)
            else:
                cur_consec = 0

        # Summary row
        pass_count = sim.eval_pass_count
        fail_count = sim.eval_fail_count
        total = pass_count + fail_count
        incomplete = getattr(sim, 'eval_incomplete_count', 0)

        # Avg profit % on passing windows, avg DD % on failing windows
        pass_profits = [r.eval_profit_pct for r in individual if r.eval_outcome == 'PASS']
        fail_dds = [r.eval_max_dd_pct for r in individual
                    if r.eval_outcome in ('FAIL_DD', 'FAIL_DAILY_DD')]

        summary_rows.append({
            'rule': fname.replace('.json', ''),
            'rule_combo': str(rule_combo)[:40],
            'exit_config': str(exit_config)[:20],
            'tf': rule_tf,
            'n_trades': len(trades),
            'total_windows': total,
            'pass_count': pass_count,
            'fail_count': fail_count,
            'incomplete': incomplete,
            'pass_rate_pct': round(100 * sim.eval_pass_rate, 1) if total else 0,
            'fail_dd': sim.eval_fail_reasons.get('FAIL_DD', 0),
            'fail_daily_dd': sim.eval_fail_reasons.get('FAIL_DAILY_DD', 0),
            'fail_timeout': sim.eval_fail_reasons.get('FAIL_TIMEOUT', 0),
            'avg_profit_pct_passed': round(sum(pass_profits) / len(pass_profits), 2) if pass_profits else 0,
            'avg_dd_pct_failed': round(sum(fail_dds) / len(fail_dds), 2) if fail_dds else 0,
            'max_consec_fails': max_consec,
            'avg_days_to_pass': round(sim.eval_avg_days_to_pass, 1),
            # WHY (June 2026): per-rule xlsx header shows Total P&L, Worst DD and
            #      Passed count — provide those keys here so the header resolves.
            'total_pnl': round(float(rd.get('total_dollar_pnl', 0)
                               or sum(float(t.get('net_profit', 0) or 0) for t in trades)), 2),
            'worst_dd_pct': round(max((r.eval_max_dd_pct for r in individual), default=0), 2),
            'passed': pass_count,
        })

        # Window rows
        for r in individual:
            window_rows.append({
                'rule': fname.replace('.json', ''),
                'exit_config': str(exit_config)[:20],
                'start_date': r.start_date,
                'outcome': r.eval_outcome,
                'profit_pct': round(r.eval_profit_pct, 2),
                'max_dd_pct': round(r.eval_max_dd_pct, 2),
                'eval_days': r.eval_days,
                'eval_trading_days': r.eval_trading_days,
            })

        rate_str = f"{sim.eval_pass_rate*100:.0f}%"
        print(f"  {fname[:50]:<50s} {rate_str:>5s} ({pass_count}/{total})"
              f"  consec_fails={max_consec}")

    if not summary_rows:
        print("[EVAL] No rules produced results — skipping report")
        return None

    # ── Build XLSX ────────────────────────────────────────────────────────
    eval_dir = os.path.join(reports_dir, 'eval_windows')
    # Clean old files so previous-run leftovers don't interfere with copy
    if os.path.isdir(eval_dir):
        for _old in os.listdir(eval_dir):
            try:
                os.remove(os.path.join(eval_dir, _old))
            except Exception:
                pass
    os.makedirs(eval_dir, exist_ok=True)

    wb = Workbook()

    # Colors
    GREEN  = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    RED    = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    ORANGE = PatternFill(start_color='FFE4B5', end_color='FFE4B5', fill_type='solid')
    GREY   = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    HEADER = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    bold = Font(bold=True, size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    # ── Summary sheet ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Summary'

    # Title row
    ws.cell(row=1, column=1, value=f'Eval Windows Report — {firm_id} ${account_size:,}')
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f'Generated {datetime.now().strftime("%Y-%m-%d %H:%M")}')
    ws.cell(row=2, column=1).font = Font(italic=True, size=9, color='666666')

    headers = ['Rule', 'Combo', 'Exit', 'TF', 'Trades', 'Windows',
               'Pass', 'Fail', 'Pass%', 'Fail_DD', 'Fail_DailyDD', 'Fail_Timeout',
               'Avg Profit% (pass)', 'Avg DD% (fail)', 'Max Consec Fails',
               'Avg Days to Pass']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.fill = HEADER
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    for i, sr in enumerate(summary_rows, 5):
        vals = [sr['rule'], sr['rule_combo'], sr['exit_config'], sr['tf'],
                sr['n_trades'], sr['total_windows'], sr['pass_count'], sr['fail_count'],
                sr['pass_rate_pct'], sr['fail_dd'], sr['fail_daily_dd'], sr['fail_timeout'],
                sr['avg_profit_pct_passed'], sr['avg_dd_pct_failed'],
                sr['max_consec_fails'], sr['avg_days_to_pass']]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.border = thin_border

        # Color pass rate
        rate_cell = ws.cell(row=i, column=9)
        if sr['pass_rate_pct'] >= 70:
            rate_cell.fill = GREEN
        elif sr['pass_rate_pct'] < 40:
            rate_cell.fill = RED

        # Color max consec fails
        consec_cell = ws.cell(row=i, column=15)
        if sr['max_consec_fails'] > 3:
            consec_cell.fill = RED
            consec_cell.font = Font(bold=True, color='9C0006')
        elif sr['max_consec_fails'] <= 1:
            consec_cell.fill = GREEN

    # Auto-width
    for col in ws.columns:
        max_len = max(len(str(c.value or '')) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 30)

    # ── Windows sheet ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Windows')
    w_headers = ['Rule', 'Exit', 'Start Date', 'Outcome', 'Profit%', 'Max DD%',
                 'Eval Days', 'Trading Days']
    for c, h in enumerate(w_headers, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.fill = HEADER
        cell.font = header_font
        cell.border = thin_border

    for i, wr in enumerate(window_rows, 2):
        vals = [wr['rule'], wr['exit_config'], wr['start_date'], wr['outcome'],
                wr['profit_pct'], wr['max_dd_pct'], wr['eval_days'],
                wr['eval_trading_days']]
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(row=i, column=c, value=v)
            cell.border = thin_border

        # Color by outcome
        outcome = wr['outcome']
        fill = None
        if outcome == 'PASS':
            fill = GREEN
        elif outcome == 'FAIL_DD':
            fill = RED
        elif outcome == 'FAIL_DAILY_DD':
            fill = ORANGE
        elif outcome == 'INSUFFICIENT_TRADES':
            fill = GREY

        if fill:
            for c in range(1, len(w_headers) + 1):
                ws2.cell(row=i, column=c).fill = fill

    for col in ws2.columns:
        max_len = max(len(str(c.value or '')) for c in col)
        ws2.column_dimensions[col[0].column_letter].width = min(max_len + 3, 30)

    # Save combined report
    out_path = os.path.join(eval_dir, 'eval_windows_report.xlsx')
    wb.save(out_path)
    print(f"\n[EVAL] Report saved: {out_path}")
    print(f"[EVAL] {len(summary_rows)} rules, {len(window_rows)} windows total")

    # ── Per-rule xlsx files (one per rule+exit combo) ─────────────────────
    # WHY (June 2026): Group by (rule, exit_config) so each combo gets its
    #      own file. Named eval_{rule}_{exit}.xlsx for easy matching to EAs.
    # CHANGED: June 2026 — group by rule+exit, add P&L to header
    _grouped = {}
    for wr in window_rows:
        _key = (wr.get('rule', ''), wr.get('exit_config', ''))
        _grouped.setdefault(_key, []).append(wr)

    _summary_by_key = {}
    for sr in summary_rows:
        _key = (sr.get('rule', ''), sr.get('exit_config', ''))
        _summary_by_key[_key] = sr

    print(f"[EVAL] Per-rule grouping: {len(_grouped)} groups")
    for _key in list(_grouped.keys())[:5]:
        print(f"[EVAL]   group key: {_key!r} -> {len(_grouped[_key])} windows")
    for (_rule, _exit), windows in _grouped.items():
        if not _rule:
            print(f"[EVAL]   SKIP empty rule: key=({_rule!r}, {_exit!r})")
            continue
        print(f"[EVAL]   generating eval_{_rule}_{_exit}.xlsx ({len(windows)} windows)")
        sr = _summary_by_key.get((_rule, _exit), {})
        wb2 = Workbook()
        ws_r = wb2.active
        ws_r.title = 'Eval Windows'

        # Info header
        ws_r.cell(row=1, column=1, value='Rule').font = bold
        ws_r.cell(row=1, column=2, value=_rule)
        ws_r.cell(row=1, column=3, value='Exit').font = bold
        ws_r.cell(row=1, column=4, value=_exit)
        ws_r.cell(row=1, column=5, value='Pass Rate').font = bold
        _pr = sr.get('pass_rate', sr.get('pass_rate_pct', 0)) or 0
        ws_r.cell(row=1, column=6, value=f"{_pr}%")

        ws_r.cell(row=2, column=1, value='Combo').font = bold
        ws_r.cell(row=2, column=2, value=str(sr.get('rule_combo', ''))[:50])
        ws_r.cell(row=2, column=3, value='Total P&L').font = bold
        ws_r.cell(row=2, column=4, value=sr.get('total_pnl', 0))
        ws_r.cell(row=2, column=4).number_format = '$#,##0.00'
        ws_r.cell(row=2, column=5, value='Max Consec Fails').font = bold
        _cf = sr.get('max_consec_fails', 0) or 0
        ws_r.cell(row=2, column=6, value=_cf)
        if _cf > 3:
            ws_r.cell(row=2, column=6).fill = RED
            ws_r.cell(row=2, column=6).font = Font(bold=True, color='9C0006')

        ws_r.cell(row=3, column=1, value='Trades').font = bold
        ws_r.cell(row=3, column=2, value=sr.get('py_trades', sr.get('n_trades', 0)))
        ws_r.cell(row=3, column=3, value='Worst DD').font = bold
        ws_r.cell(row=3, column=4, value=f"{sr.get('worst_dd_pct', 0)}%")
        ws_r.cell(row=3, column=5, value=f"Passed {sr.get('passed',0)}/{sr.get('total_windows',0)}").font = bold

        # Window headers
        for c, h in enumerate(w_headers, 1):
            cell = ws_r.cell(row=4, column=c, value=h)
            cell.fill = HEADER
            cell.font = header_font
            cell.border = thin_border

        for i, wr in enumerate(windows, 5):
            vals = [wr['rule'], wr['exit_config'], wr['start_date'], wr['outcome'],
                    wr['profit_pct'], wr['max_dd_pct'], wr['eval_days'],
                    wr['eval_trading_days']]
            for c, v in enumerate(vals, 1):
                cell = ws_r.cell(row=i, column=c, value=v)
                cell.border = thin_border

            outcome = wr['outcome']
            fill = None
            if outcome == 'PASS':
                fill = GREEN
            elif outcome == 'FAIL_DD':
                fill = RED
            elif outcome == 'FAIL_DAILY_DD':
                fill = ORANGE
            elif outcome == 'INSUFFICIENT_TRADES':
                fill = GREY
            if fill:
                for c in range(1, len(w_headers) + 1):
                    ws_r.cell(row=i, column=c).fill = fill

        safe_name = f"eval_{_rule}_{_exit}".replace('/', '_').replace('\\', '_')[:80]
        per_path = os.path.join(eval_dir, f'{safe_name}.xlsx')
        wb2.save(per_path)

    print(f"[EVAL] {len(_grouped)} per-rule xlsx files saved to {eval_dir}/")
    return out_path


if __name__ == '__main__':
    import argparse
    p = argparse.ArgumentParser(description='Compare MT5 reports folder to Python trades')
    p.add_argument('--reports', required=True, help='folder of MT5 .xlsx reports')
    p.add_argument('--python', required=True,
                   help='folder of Python trade CSVs (per-rule or a shared trades.csv)')
    p.add_argument('--manifest', default='', help='batch_manifest.json (optional labels)')
    p.add_argument('--bar-minutes', type=int, default=5, help='bar size for shift check')
    args = p.parse_args()
    compare_folder(args.reports, args.python,
                   manifest_path=args.manifest or None,
                   bar_minutes=args.bar_minutes)
