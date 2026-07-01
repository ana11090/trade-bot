# WHY: expose batch EA generation + MT5 run-file emit + report compare in one panel.
# CHANGED: June 2026 — new panel; reuses batch_ea_tools + batch_compare_reports.
# CHANGED: June 2026 — added checkbox rule grid (same interaction as Strategy Refiner)
# CHANGED: June 2026 — grid columns + filters match Strategy Refiner grid
# CHANGED: June 2026 — load via load_strategy_list() so backtest-matrix rows (with stats) appear
import os
import threading
import tkinter as tk
from tkinter import ttk, filedialog

from shared.my_rules import load_all as _load_my_rules
try:
    from shared.saved_rules import load_all as _load_saved_rules
except Exception:
    _load_saved_rules = None

# WHY: load_strategy_list() returns backtest-matrix rows (with full stats) merged with saved/my
#   rules — the same source the refiner grid uses. Without this the batch grid only sees raw
#   saved/my_rules which are discovery-only and have no trade stats.
# CHANGED: June 2026 — load via strategy_refiner.load_strategy_list
try:
    from project2_backtesting.strategy_refiner import load_strategy_list as _load_strategy_list
except Exception:
    _load_strategy_list = None

# WHY: reuse the refiner's display formatters so money/win-pass/prop-score formatting
#   matches exactly across panels.
# CHANGED: June 2026 — batch grid parity with refiner grid
try:
    from project2_backtesting.panels.strategy_refiner_panel import (
        _money_for_strategy, _format_win_pass, _format_prop_score, _compute_prop_score,
    )
except Exception:
    _money_for_strategy  = None
    _compute_prop_score  = None
    _format_win_pass     = lambda s: "—"
    _format_prop_score   = lambda s: "—"

BG      = "#f0f2f5"
DARK    = "#1a1a2a"
MIDGREY = "#555566"
WHITE   = "#ffffff"

# CHANGED: June 2026 — checkbox grid state
_log            = None
_grid_tree      = None      # the Treeview
_grid_entries   = []        # strategy dicts (flat, from load_strategy_list)
_batch_sel_iids = set()     # iids (str) currently ticked
# WHY: iid is now r.get('id') or a counter — not a positional index — so we need
#   a direct iid→entry map for _do_generate to look up selected rows safely.
# CHANGED: June 2026 — iid_to_entry avoids int(iid)-as-index assumption
_iid_to_entry   = {}        # iid_str -> strategy dict
# Filter widgets/vars — set in build_panel as module globals, read by _populate_grid.
# WHY: must be module-level (not build_panel locals) so _populate_grid can write the
#   global declaration and read them reliably regardless of call site.
# CHANGED: June 2026 — promoted to module globals; scope fix
_f_stage      = None
_f_tf         = None
_f_dir        = None
_f_mintr      = None
_f_minwr      = None
_f_sort       = None
_f_profitable = None   # BooleanVar — profitable-only filter (parity with refiner)
_nostats_lbl  = None
_cur_source   = "all"  # tracks the active source so filter-widget callbacks don't need src_var
_latest_run_at = None  # generated_at of the most recently observed backtest matrix (for "this run" filter)
# CHANGED: June 2026 — Run Backtest / Run Scenario mode toggle
# WHY: "Run Scenario" loads the SAME freshly-discovered rules the Run Backtest
#   panel loads (via shared.scenario_sources), straight to MT5 — no Python
#   backtest in between (the batch pipeline never ran one anyway).
_mode_var       = None   # "Run Backtest" | "Run Scenario"
_scenario_var   = None   # selected scenario-source label
_scenario_paths = {}     # {label: path} from shared.scenario_sources
_scenario_combo = None   # the scenario dropdown widget (hidden in backtest mode)
_src_combo_ref  = None   # the existing Source dropdown widget (hidden in scenario mode)
# CHANGED: June 2026 — account-size dropdown (rule's own account wins; this is the
#   fallback). Applies in BOTH modes; populated from the firm's account_sizes.
_acct_var       = None   # selected account size (str), e.g. "10000"
# CHANGED: June 2026 — multi-select exit-strategy filter
_f_exit_vars  = {}     # exit_name -> tk.BooleanVar
_f_exit_all   = None   # tk.BooleanVar for the "All" toggle
_f_exit_menu  = None   # the Menubutton (so we can rebuild its menu when rows change)
_f_exit_menuobj = None # the tk.Menu object itself (NOT _f_exit_menu['menu'], which is a str)


# CHANGED: June 2026 — helper returns checked exits or None for 'all'
# WHY: exit filter logic needs a stable query function; returns None when "All" is checked.
def _selected_exits():
    """Return set of checked exit names, or None if "All" is checked (or nothing is checked)."""
    global _f_exit_all, _f_exit_vars
    if _f_exit_all is not None and _f_exit_all.get():
        return None  # All → no filtering
    picked = {n for n, v in _f_exit_vars.items() if v.get()}
    return picked or None  # If nothing is picked, treat as "All"


# CHANGED: June 2026 — rebuild exit filter menu from current grid rows
# WHY: each source has different exit strategies; rebuild the checkboxes dynamically.
# HOTFIX: June 2026 — use _f_exit_menuobj (the real Menu object), not _f_exit_menu['menu'] (a str)
# FIX: June 2026 — proper checkbuttons with independent BooleanVars; menu stays open on toggle.
def _rebuild_exit_menu(all_rows=None):
    """Scan rows for unique exits, rebuild the Menubutton's menu with checkboxes."""
    global _f_exit_vars, _f_exit_all, _f_exit_menu, _f_exit_menuobj
    # WHY: panel not built yet → menu object doesn't exist; skip safely to prevent blank panel.
    m = _f_exit_menuobj  # the real Menu object, never _f_exit_menu['menu'] (which is a str)
    if m is None:
        return

    # Extract unique exit names from rows (all_rows if provided, else _grid_entries)
    rows = all_rows if all_rows is not None else _grid_entries
    names = sorted({
        (r.get('exit_name') or r.get('exit_class') or
         r.get('exit_strategy') or '').strip()
        for r in rows
        if (r.get('exit_name') or r.get('exit_class') or r.get('exit_strategy'))
    })

    # Clear old menu
    m.delete(0, 'end')

    # "All" as a CHECKBUTTON (keeps menu open, shows a check)
    if _f_exit_all is None:
        _f_exit_all = tk.BooleanVar(value=True)

    def _on_all():
        """When All is toggled, clear individual picks (or re-enable All if nothing else is checked)."""
        if _f_exit_all.get():
            # Turning All on clears individual picks
            for v in _f_exit_vars.values():
                v.set(False)
        else:
            # Don't allow All to be the only thing unchecked with nothing else → re-check it
            if not any(v.get() for v in _f_exit_vars.values()):
                _f_exit_all.set(True)
        _refresh_exit_label()
        _populate_grid(_cur_source)

    m.add_checkbutton(label="All", variable=_f_exit_all, command=_on_all)
    m.add_separator()

    # One CHECKBUTTON per exit
    for nm in names:
        if nm not in _f_exit_vars:
            _f_exit_vars[nm] = tk.BooleanVar(value=False)

        def _on_one(_nm=nm):
            """When a specific exit is toggled, turn All off (or back on if nothing is checked)."""
            if any(v.get() for v in _f_exit_vars.values()):
                _f_exit_all.set(False)  # Picking specific exits turns All off
            else:
                _f_exit_all.set(True)   # Nothing left → back to All
            _refresh_exit_label()
            _populate_grid(_cur_source)

        m.add_checkbutton(label=nm, variable=_f_exit_vars[nm], command=_on_one)

    # Drop stale vars (exits that no longer exist in the current source)
    for stale in [k for k in list(_f_exit_vars) if k not in names]:
        _f_exit_vars.pop(stale, None)

    _refresh_exit_label()


# CHANGED: June 2026 — update Menubutton label to show current selection
# WHY: "All ▾" / "PSAR Only ▾" / "3 exits ▾" tells you what's filtered without opening menu.
def _refresh_exit_label():
    """Update the Menubutton text to reflect current selection."""
    global _f_exit_all, _f_exit_vars, _f_exit_menu
    if _f_exit_menu is None:
        return
    # If "All" is checked, show "All ▾"
    if _f_exit_all is None or _f_exit_all.get():
        _f_exit_menu.config(text="All ▾")
        return
    # Otherwise, show the specific exits that are checked
    picked = [n for n, v in _f_exit_vars.items() if v.get()]
    if not picked:
        _f_exit_menu.config(text="All ▾")
    elif len(picked) == 1:
        # Truncate long names to 14 chars
        _f_exit_menu.config(text=picked[0][:14] + " ▾")
    else:
        _f_exit_menu.config(text="%d exits ▾" % len(picked))


# CHANGED: June 2026 — remember last generate out_dir so 1b/2 default to it
_last_out_dir    = None
# CHANGED: June 2026 — remember run artifacts so "Run Tests" / Compare need no prompts
_last_bat_path   = None
_last_reports_dir = None
# CHANGED: June 2026 — persist the MT5 data dir per session so canonical_batch_dir never re-asks
_last_data_dir   = None


# CHANGED: June 2026 — derive MT5 data dir: folder above the FIRST 'MQL5' segment.
#   Idempotent: works even on doubled paths like ...MQL5\Experts\batch\MQL5\Experts\batch.
def _derive_data_dir(path):
    """Return the MT5 data dir (folder directly above the first 'MQL5' segment).
    If no MQL5 segment exists, treat the path itself as the data dir."""
    if not path:
        return None
    p = os.path.abspath(path)
    parts = p.replace('/', '\\').split('\\')
    for i, seg in enumerate(parts):
        if seg.lower() == 'mql5':
            if i == 0:
                return None
            return '\\'.join(parts[:i])   # everything before the first MQL5
    return p   # no MQL5 found — treat as data dir


# CHANGED: June 2026 — single source of truth for the batch EA folder. All steps (generate,
#   compile, build-run-files, run) call this so the .ini's Expert=batch\name.ex5 always resolves.
#   Self-correcting: collapses any doubled path back to the true data dir before appending once.
def _canonical_batch_dir():
    """Return <data_dir>\\MQL5\\Experts\\batch. Never doubles the path."""
    global _last_data_dir
    # Derive data dir from whatever we know, collapsing any nesting via _derive_data_dir.
    dd = _last_data_dir or (_derive_data_dir(_last_out_dir) if _last_out_dir else None)
    if not dd or not os.path.isdir(dd):
        picked = filedialog.askdirectory(
            title="Pick MT5 DATA folder (the long-hash folder containing MQL5\\Experts)")
        if not picked:
            return None
        # Collapse whatever they picked back to the real data dir (handles picking MQL5 or deeper)
        dd = _derive_data_dir(picked) or picked
    _last_data_dir = dd
    return os.path.join(dd, "MQL5", "Experts", "batch")


# CHANGED: June 2026 — detect a running MT5 terminal before launching the tester bat
def _mt5_is_running():
    try:
        import subprocess as _sp
        out = _sp.run(["tasklist", "/FI", "IMAGENAME eq terminal64.exe"],
                      capture_output=True, text=True, timeout=10)
        return "terminal64.exe" in (out.stdout or "")
    except Exception:
        return False  # if we can't tell, don't block


# ── grid helpers ──────────────────────────────────────────────────────────────

# WHY: matrix net/pf/etc. can arrive as str or non-native numeric. The refiner coerces with
#   float() (strategy_refiner_panel.py:6498) — the strict isinstance() gate was rejecting
#   those values, so losing rows never got the ❌/losing tag and "Profitable only" had nothing
#   to filter. Coerce the same way the refiner does.
# CHANGED: June 2026 — tolerant numeric coercion (handles str like "1,234" / "-1234")
def _as_num(v):
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace(',', '').replace('+', '').replace('−', '-')
        try:
            return float(s)
        except ValueError:
            return None
    return None

# CHANGED: June 2026 — stage cell; strategy dicts from load_strategy_list are flat
def _stage_cell(r):
    # WHY: prop_firm_stage lives at the top level of load_strategy_list() dicts (flat).
    #   For saved/my_rules paths it may also be in a nested 'saved_rule' sub-dict.
    v = r.get('prop_firm_stage')
    if v:
        return str(v)
    saved = r.get('saved_rule') or {}
    v = saved.get('prop_firm_stage')
    if v:
        return str(v)
    rs = r.get('run_settings') or {}
    v = rs.get('prop_firm_stage')
    if v:
        return str(v)
    # WHY (fix 1B): backtest rows predate or omit the stage stamp; infer "evaluation"
    #   when the row has firm info rather than showing "—". Cosmetic — the firm phase
    #   dropdown drives actual pass-rate numbers, not this column.
    # CHANGED: July 2026 — evaluation fallback for unstamped backtest rows
    if r.get('firm_id') or r.get('prop_firm_name'):
        return "evaluation"
    return "—"


def _rule_label(r):
    # WHY: strategy dicts from load_strategy_list() carry rule_combo at the top level;
    #   saved rows may use rule_label, name, or id. Tolerant — never crashes.
    # CHANGED: June 2026 — wider fallback chain; saved rows vary
    return (r.get('rule_combo') or r.get('rule') or r.get('rule_label') or
            r.get('label') or r.get('name') or str(r.get('id', '—')))


def _populate_grid(source):
    # WHY: reload the grid from load_strategy_list() (same merged source the refiner uses),
    #   filter by source tag, apply UI filters, sort, then insert.
    # CHANGED: June 2026 — load_strategy_list + filter by source tag + refiner-parity columns
    # CHANGED: June 2026 — defensive per-row try/except so one bad row can't blank the grid;
    #   stats sub-dict fallback for saved rows; non-empty fallback when source yields 0 rows
    # CHANGED: June 2026 — scope fix: declare filter globals explicitly so _passes always sees
    #   the live widget values; filter applied as list comprehension before insert loop
    # CHANGED: June 2026 — added exit filter globals
    global _grid_entries, _batch_sel_iids, _iid_to_entry, _cur_source, _latest_run_at
    global _f_stage, _f_tf, _f_dir, _f_mintr, _f_minwr, _f_sort, _f_profitable
    global _f_exit_vars, _f_exit_all, _f_exit_menu, _f_exit_menuobj
    if _grid_tree is None:
        return
    # WHY: in Run Scenario mode every refresh/filter/TF-filter callback fires
    #   _populate_grid(_cur_source) with _cur_source="scenario" — which has no
    #   matching source in load_strategy_list() and would fall through to 'all',
    #   leaking backtest rows into the scenario grid. Delegate to the scenario
    #   renderer instead (it honors the same _f_tf filter for the 5-TF fan).
    # CHANGED: June 2026 — scenario-mode delegation guard
    if _mode_var is not None and _mode_var.get() == "Run Scenario":
        _populate_grid_scenario()
        return
    _cur_source = source
    for _i in _grid_tree.get_children():
        _grid_tree.delete(_i)
    _batch_sel_iids = set()
    _iid_to_entry = {}

    # ── diagnostic: confirm import worked and how many rows we have ──
    print("[BATCH-GRID] _load_strategy_list =", _load_strategy_list,
          " loaded =", (len(_load_strategy_list()) if _load_strategy_list else 'n/a'),
          flush=True)

    # ── load merged strategy list ──
    if _load_strategy_list is not None:
        _all = _load_strategy_list() or []
    elif _load_saved_rules:
        _all = _load_saved_rules() or []
    else:
        _all = []

    # WHY: separator rows are section dividers with no id/rule/stats — they crash the insert
    #   loop if they reach it. Strip them unconditionally before any further processing.
    _all = [r for r in _all if r.get('source') != 'separator']

    # Pre-compute backtest rows once; track their latest generated_at for "this run".
    # CHANGED: July 2026 — "this run" scoping (fix issue 2A)
    _bt_all = [r for r in _all if r.get('source') == 'backtest']
    if _bt_all:
        _lat = max((r.get('generated_at', '') for r in _bt_all), default=None)
        if _lat:
            _latest_run_at = _lat

    # Filter by the source dropdown.
    # Source tags set by load_strategy_list: 'backtest', 'optimizer', 'saved', 'my_rules'.
    # WHY: 'saved_rules' catches everything that isn't backtest/optimizer (source values from
    #   saved entries vary: 'saved', 'my_rules', '?', or sometimes absent).
    if source == 'backtest':
        _grid_entries = list(_bt_all)
    elif source == 'this run':
        # WHY: show only rows from the most recently written backtest matrix.
        #   generated_at is propagated from the matrix JSON top-level to each row
        #   by load_strategy_list(). Rows from an older run have a lower timestamp
        #   and are excluded. Falls back to all backtest rows when no timestamp.
        # CHANGED: July 2026 — "this run" filter
        if _latest_run_at:
            _grid_entries = [r for r in _bt_all if r.get('generated_at') == _latest_run_at]
        else:
            _grid_entries = list(_bt_all)
    elif source == 'optimizer':
        _grid_entries = [r for r in _all if r.get('source') == 'optimizer']
    elif source == 'saved_rules':
        _grid_entries = [r for r in _all
                         if r.get('source') not in ('backtest', 'optimizer')]
    elif source == 'my_rules':
        _grid_entries = [r for r in _all if r.get('source') == 'my_rules']
    else:  # 'all'
        _grid_entries = list(_all)
        # WHY: load_strategy_list() can return the same rule id under multiple sources
        #   (e.g. backtest copy id=41 AND its saved copy id=41). De-dupe by id, keeping
        #   the copy with the best source (backtest > optimizer > saved > my_rules) so
        #   the row with real stats wins and 'all' always has ≥ rows as any single source.
        # CHANGED: June 2026 — de-dupe 'all' to prevent iid collision and show each rule once
        _src_rank = {'backtest': 0, 'optimizer': 1, 'saved': 2, 'my_rules': 3}
        _best = {}
        for _r in _grid_entries:
            _rid = str(_r.get('id', id(_r)))
            _cur = _best.get(_rid)
            if (_cur is None or
                    _src_rank.get(_r.get('source'), 9) < _src_rank.get(_cur.get('source'), 9)):
                _best[_rid] = _r
        _grid_entries = list(_best.values())

    # WHY: if the chosen source is empty (e.g. backtest_matrix.json is a Git LFS pointer /
    #   not pulled), fall back to 'all' so the page is never blank when data exists elsewhere.
    # CHANGED: June 2026 — non-empty fallback
    if not _grid_entries and _all:
        print("[BATCH-GRID] source '%s' empty — falling back to all" % source, flush=True)
        _grid_entries = list(_all)

    # CHANGED: June 2026 — rebuild exit filter menu based on current rows (before filtering)
    # WHY: each source has different exit strategies; dynamically populate the checkboxes.
    # HOTFIX: June 2026 — wrap in try/except so exit-menu issues can't blank the whole panel.
    try:
        _rebuild_exit_menu()
    except Exception as _e:
        print("[BATCH-GRID] exit menu rebuild skipped:", _e, flush=True)

    # ── helpers used by filter and sort ──
    def _val(var, default=""):
        # Read a Tk variable/widget safely; return default if None or widget destroyed.
        try:
            return var.get() if var is not None else default
        except Exception:
            return default

    def _num(r, *keys):
        # CHANGED: June 2026 — coerce via _as_num (str/non-native ok), matches refiner behaviour
        stats = r.get('stats') or {}
        for k in keys:
            raw = r.get(k) if r.get(k) is not None else stats.get(k)
            val = _as_num(raw)
            if val is not None:
                return val
        return None

    # ── filter predicate — applied to _grid_entries before the insert loop ──
    # WHY: filtering the list before insert (not per-row in the loop) means the
    #   insert loop never skips rows mid-flight, making _shown count accurate and
    #   the iid mapping stable.
    # CHANGED: June 2026 — filter applied as list comprehension; _val() guards None vars
    # CHANGED: June 2026 — added exit-strategy filter (multi-select)
    def _passes(r):
        st = _stage_cell(r)
        tf = r.get('entry_tf') or r.get('entry_timeframe') or '—'
        dr = r.get('direction') or r.get('dir') or '—'
        # CHANGED: June 2026 — strip whitespace from exit name for consistent matching
        exit_ = (r.get('exit_name') or r.get('exit_class') or
                 r.get('exit_strategy') or '').strip()
        tr = _num(r, 'total_trades') or 0
        wr = _num(r, 'win_rate') or 0
        net = _num(r, 'net_pips', 'net_total_pips', 'total_pips')

        if _val(_f_stage, "All") not in ("All", "", st):
            return False
        if _val(_f_tf, "All") not in ("All", "", tf):
            return False
        if _val(_f_dir, "All") not in ("All", "", dr):
            return False
        # CHANGED: June 2026 — exit filter: if specific exits are checked, hide others
        # WHY: _selected_exits() returns None when "All" is on, set of names otherwise.
        _sel_exits = _selected_exits()
        if _sel_exits is not None and exit_ and exit_ not in _sel_exits:
            return False
        try:
            if _val(_f_mintr) and tr < float(_val(_f_mintr)):
                return False
            if _val(_f_minwr) and wr < float(_val(_f_minwr)):
                return False
        except ValueError:
            pass
        # CHANGED: June 2026 — profitable-only filter; matches refiner: hides rules that HAVE
        #   data and are ≤ 0; keeps no-data rows visible (same guard as refiner uses)
        if _val(_f_profitable, False):
            if net is not None and net <= 0:
                return False
        return True

    _grid_entries = [r for r in _grid_entries if _passes(r)]

    # ── sort ──
    # CHANGED: June 2026 — sort after filter; field names match matrix rows (flat, top-level)
    _sort_choice = _val(_f_sort, "Net Pips")
    _sort_map = {
        "Prop Score": "prop_score",
        "Net Pips":   "net_total_pips",
        "Win Rate":   "win_rate",
        "PF":         "net_profit_factor",
        "Trades":     "total_trades",
    }
    _sk = _sort_map.get(_sort_choice, "net_total_pips")

    def _sort_key(r):
        if _sort_choice == "Prop Score" and _compute_prop_score is not None:
            try:
                score, _ = _compute_prop_score(r)
                return score or 0
            except Exception:
                return 0
        # CHANGED: June 2026 — _num already coerces via _as_num; string vals sort correctly
        return _num(r, _sk) or 0.0

    _grid_entries.sort(key=_sort_key, reverse=True)

    # ── insert rows ──
    # WHY: strategy dicts from load_strategy_list can be flat (backtest rows) or have a
    #   'stats' sub-dict (some saved rows). Read flat first, fall back to sub-dict.
    #   Each row is wrapped in try/except so one bad entry can never blank the whole grid.
    # CHANGED: June 2026 — per-row guard; iid = id/counter stored in _iid_to_entry
    # CHANGED: June 2026 — safe number formatter (_fmt) replaces inline :+,.Nf specs which
    #   crash when Python sees the ',' flag combined with '+' in some value paths.
    def _fmt(v, dec=1, sign=False, comma=False):
        if not isinstance(v, (int, float)):
            return "—"
        try:
            s = ("%.{d}f".replace("{d}", str(dec)) % v) if not comma else (
                "{:,.{d}f}".format(v, d=dec))
            if sign and v >= 0:
                s = "+" + s
            elif sign:
                pass  # negative sign already in s
            return s
        except (ValueError, TypeError):
            return "—"

    _no_stats = 0
    _shown = 0
    for r in _grid_entries:
        # separators were stripped earlier; this is a belt-and-braces guard only
        if r.get('source') == 'separator':
            continue
        try:
            stats  = r.get('stats') or {}
            # CHANGED: June 2026 — coerce all numeric fields via _as_num so string values
            #   (e.g. "1,234" stored in matrix JSON) don't produce None and blank the cells.
            trades = _as_num(r.get('total_trades') if r.get('total_trades') is not None else stats.get('total_trades'))
            wr     = _as_num(r.get('win_rate')     if r.get('win_rate')     is not None else stats.get('win_rate'))
            pf     = _as_num(r.get('net_profit_factor') or r.get('profit_factor') or
                             stats.get('net_profit_factor') or stats.get('profit_factor'))
            net    = _as_num(r.get('net_total_pips') or r.get('net_pips') or
                             stats.get('net_total_pips') or stats.get('net_pips'))
            avg    = _as_num(r.get('net_avg_pips') or r.get('avg_pips') or
                             stats.get('net_avg_pips') or stats.get('avg_pips'))
            tf     = (r.get('entry_tf') or r.get('entry_timeframe') or
                      stats.get('entry_tf') or '—')
            exit_  = (r.get('exit_name') or r.get('exit_class') or
                      r.get('exit_strategy') or stats.get('exit_name') or '?')
            direction = r.get('direction') or r.get('dir') or '—'
            try:
                off = "N" if int(r.get('entry_bar_offset', 0) or 0) == 0 else "N+1"
            except (TypeError, ValueError):
                off = "N"

            has_stats = trades is not None
            if not has_stats:
                _no_stats += 1

            # CHANGED: June 2026 — coerce net (matches refiner net>0 definition); net_total_pips
            #   listed first (same priority as refiner line 6571). _as_num handles string values.
            _nv = None
            for _nk in ('net_total_pips', 'net_pips', 'total_pips'):
                _nv = _as_num(stats.get(_nk) if stats.get(_nk) is not None else r.get(_nk))
                if _nv is not None:
                    break
            if not has_stats or _nv is None:
                profit_cell = "—";  profit_tag = "neutral"
            elif _nv > 0:
                profit_cell = "✅"; profit_tag = "profitable"
            else:
                profit_cell = "❌"; profit_tag = "losing"

            if has_stats and _money_for_strategy is not None:
                _nd, _np, _ad, _ap = _money_for_strategy(r, net or 0, avg or 0)
                net_d = ("$" + _fmt(_nd, 0, sign=True, comma=True)) if _nd is not None else "—"
                net_p = (_fmt(_np, 1, sign=True) + "%")             if _np is not None else "—"
                avg_d = ("$" + _fmt(_ad, 2, sign=True, comma=True)) if _ad is not None else "—"
                avg_p = (_fmt(_ap, 2, sign=True) + "%")             if _ap is not None else "—"
            else:
                net_d = net_p = avg_d = avg_p = "—"

            vals = (
                "☐",
                r.get('id', r.get('index', _shown)),
                _stage_cell(r),
                str(_rule_label(r))[:60],
                exit_, tf, direction,
                (int(trades) if has_stats else "—"),
                (_fmt(wr, 1) + "%") if (has_stats and wr is not None) else "—",
                _fmt(pf, 2)         if (has_stats and pf is not None) else "—",
                _fmt(net, 0, sign=True, comma=True) if (has_stats and net is not None) else "—",
                net_d, net_p,
                _fmt(avg, 1, sign=True) if (has_stats and avg is not None) else "—",
                avg_d, avg_p,
                profit_cell,
                (_format_win_pass(r)   if has_stats else "—"),
                (_format_prop_score(r) if has_stats else "—"),
                off,
            )
            # WHY: the same rule id appears under multiple sources (backtest + saved copy).
            #   Treeview.insert raises TclError on a duplicate iid — swallowed by the except
            #   and silently dropped rows, making 'all' < 'backtest'. Unique iid per row.
            # CHANGED: June 2026 — unique iid = source::id::counter (no TclError collisions)
            _src = str(r.get('source', 'x'))
            _rid = str(r.get('id', r.get('index', _shown)))
            _iid = "%s::%s::%d" % (_src, _rid, _shown)
            _iid_to_entry[_iid] = r
            _grid_tree.insert('', 'end', iid=_iid, values=vals, tags=(profit_tag,))
            _shown += 1
        except Exception as _row_err:
            print("[BATCH-GRID] skipped a row:", repr(_row_err), flush=True)
            continue

    print("[BATCH-GRID] source=%s shown=%d no_stats=%d" % (source, _shown, _no_stats),
          flush=True)
    _grid_tree.heading("sel", text="☐", command=_toggle_all)

    # ── caveat label ──
    if _nostats_lbl is not None:
        if _no_stats:
            _nostats_lbl.config(
                text="(%d of %d rules have no backtest data — re-run backtest to fill stats)"
                     % (_no_stats, _shown))
        else:
            _nostats_lbl.config(text="")


def _scenario_label(r):
    """Descriptive fallback label for a discovery rule with no id."""
    import hashlib
    _dir = r.get('direction', r.get('action', 'BUY'))
    _tf  = r.get('entry_timeframe', r.get('entry_tf', 'XX'))
    _nc  = len(r.get('conditions', []))
    _h   = hashlib.md5(str(sorted(str(c) for c in r.get('conditions', []))).encode()).hexdigest()[:4]
    return f"{_dir}_{_tf}_{_nc}c_{_h}"


def _populate_grid_scenario():
    """Populate the grid from the selected scenario source — the SAME rules the
    Run Backtest panel would load (via shared.scenario_sources). These feed
    Generate EAs -> Compile -> Build -> Run Tests -> Compare straight to MT5;
    the batch pipeline never ran a Python backtest, so no flow change is needed.

    Discovery rules carry no backtest stats, so every stats cell renders "—" —
    byte-identical to what _populate_grid produces for a stat-less row. The
    column order here MUST match `cols`/`_col_spec` in build_panel (~line 2485)
    and the vals tuple in _populate_grid (~line 517); keep all three in sync."""
    global _grid_entries, _batch_sel_iids, _iid_to_entry, _cur_source
    if _grid_tree is None:
        return
    try:
        from shared.scenario_sources import load_rules_for_source
    except Exception as _imp_err:
        print("[BATCH-GRID] scenario import failed:", repr(_imp_err), flush=True)
        return
    label = _scenario_var.get() if _scenario_var is not None else ""
    rules = load_rules_for_source(label, _scenario_paths) if label else []

    for _i in _grid_tree.get_children():
        _grid_tree.delete(_i)
    _batch_sel_iids = set()
    _iid_to_entry = {}
    _cur_source = "scenario"

    # Expand each source rule into (direction × TF × exit) variants using the
    # backtest's OWN exit set + direction-expansion + label logic (single source
    # of truth — see shared/scenario_expand.py). No Python backtest is run; these
    # variants are fed to Generate -> generate_ea (MT5). Conditions/feature
    # prefixes are unchanged — the backtester doesn't remap them either.
    from shared.scenario_expand import expand_scenario_rules
    _fanned = expand_scenario_rules(rules)

    # Optional TF filter (shared _f_tf combobox): narrow the fan to one TF so the
    # grid isn't overwhelming. "All" (default) keeps the full 5× fan.
    try:
        _sel_tf = _f_tf.get() if _f_tf is not None else "All"
    except Exception:
        _sel_tf = "All"
    if _sel_tf and _sel_tf.lower() != "all":
        _fanned = [_r for _r in _fanned
                   if str(_r.get("entry_tf", "")).lower() == _sel_tf.lower()]

    # CHANGED 2026-06-26 — guard against runaway fans. "All Sources Combined"
    # (~9,300 rules) × exits × TFs is ~1M variants and freezes the Treeview.
    # Cap the grid; the user narrows with the source picker + TF/exit filters.
    _MAX_SCENARIO_ROWS = 2000
    _total_variants = len(_fanned)
    _truncated = False
    if _total_variants > _MAX_SCENARIO_ROWS:
        _fanned = _fanned[:_MAX_SCENARIO_ROWS]
        _truncated = True

    _grid_entries = _fanned

    if not _grid_entries:
        # sel/id/stage blank, message in the 'rule' column; Treeview pads the rest.
        _grid_tree.insert('', 'end', iid='__none__',
                          values=("", "", "", "(no rules in this scenario source)"))
        if _nostats_lbl is not None:
            _nostats_lbl.config(text="")
        return

    _shown = 0
    for r in _grid_entries:
        try:
            _name = (r.get('_fanned_label') or r.get('_saved_rule_id') or
                     r.get('rule_id') or _scenario_label(r))
            tf     = (r.get('entry_tf') or r.get('entry_timeframe') or '—')
            exit_  = (r.get('exit_name') or r.get('exit_class') or
                      r.get('exit_strategy') or '?')
            direction = r.get('direction') or r.get('dir') or r.get('action') or '—'
            try:
                off = "N" if int(r.get('entry_bar_offset', 0) or 0) == 0 else "N+1"
            except (TypeError, ValueError):
                off = "N"
            # 20 columns: sel, id, stage, rule, exit, tf, dir, trades, wr, pf,
            #   net_pips, net_$, net_%, avg_pips, avg_$, avg_%, profit, win_pass,
            #   prop_score, off. Discovery rules have no stats -> all "—".
            vals = (
                "☐",
                r.get('id', _shown),
                _stage_cell(r),
                str(_name)[:60],
                exit_, tf, direction,
                "—",                      # trades
                "—",                      # win rate
                "—",                      # PF
                "—",                      # net pips
                "—", "—",                 # net $, net %
                "—",                      # avg pips
                "—", "—",                 # avg $, avg %
                "—",                      # profit
                "—",                      # win pass
                "—",                      # prop score
                off,
            )
            _iid = "scn::%d" % _shown
            _iid_to_entry[_iid] = r
            _grid_tree.insert('', 'end', iid=_iid, values=vals, tags=("neutral",))
            _shown += 1
        except Exception as _row_err:
            print("[BATCH-GRID] scenario skipped a row:", repr(_row_err), flush=True)
            continue

    print("[BATCH-GRID] scenario source=%r shown=%d" % (label, _shown), flush=True)
    _grid_tree.heading("sel", text="☐", command=_toggle_all)
    if _nostats_lbl is not None:
        if _truncated:
            _nostats_lbl.config(
                text=("scenario mode — showing %d of %d variants (capped). "
                      "Pick a single source or use TF/Exit filters to narrow."
                      % (_shown, _total_variants)))
        else:
            _nostats_lbl.config(
                text=("scenario mode — %d rule(s) from discovery source; "
                      "stats filled by the MT5 run" % _shown))


def _toggle_row(event):
    # WHY: column-#1 click toggles one row's tick, matching the Refiner pattern.
    if _grid_tree is None:
        return
    if _grid_tree.identify_region(event.x, event.y) != 'cell':
        return
    if _grid_tree.identify_column(event.x) != '#1':
        return
    iid = _grid_tree.identify_row(event.y)
    if not iid:
        return
    if iid in _batch_sel_iids:
        _batch_sel_iids.discard(iid)
        glyph = "☐"
    else:
        _batch_sel_iids.add(iid)
        glyph = "☑"
    vals = list(_grid_tree.item(iid, 'values'))
    if vals:
        vals[0] = glyph
        _grid_tree.item(iid, values=vals)
    return "break"


def _toggle_all():
    # WHY: header click selects or deselects every visible row at once.
    if _grid_tree is None:
        return
    kids = list(_grid_tree.get_children())
    all_on = bool(kids) and all(k in _batch_sel_iids for k in kids)
    glyph = "☐" if all_on else "☑"
    for k in kids:
        if all_on:
            _batch_sel_iids.discard(k)
        else:
            _batch_sel_iids.add(k)
        vals = list(_grid_tree.item(k, 'values'))
        if vals:
            vals[0] = glyph
            _grid_tree.item(k, values=vals)
    _grid_tree.heading("sel", text=("☑" if not all_on else "☐"), command=_toggle_all)


# ── log helper ────────────────────────────────────────────────────────────────

def _append(msg):
    # WHY: thread-safe append; marshal via after() since Tk is single-threaded.
    if _log is None:
        return
    def _do():
        _log.configure(state="normal")
        _log.insert("end", msg + "\n")
        _log.see("end")
        _log.configure(state="disabled")
    _log.after(0, _do)


def _run_bg(fn):
    # WHY: keep the UI responsive during slow operations.
    threading.Thread(target=fn, daemon=True).start()


# ── button callbacks ──────────────────────────────────────────────────────────

# CHANGED: June 2026 — synchronous step functions for "Run All" chain
# WHY: each step returns True/False so the chain can stop on failure.
def _step_generate(source_var):
    """Synchronous: returns True on success, False on error."""
    try:
        from project3_live_trading.batch_ea_tools import batch_generate
        _selected = [_iid_to_entry[i] for i in _batch_sel_iids if i in _iid_to_entry]
        if not _selected:
            _append("Tick at least one rule in the grid (or the header ☐ to select all).")
            return False
        out_dir = _canonical_batch_dir()
        if not out_dir:
            _append("Generate cancelled (no MT5 data folder).")
            return False
        os.makedirs(out_dir, exist_ok=True)
        global _last_out_dir
        _last_out_dir = out_dir
        _nested = os.path.join(out_dir, "batch")
        if os.path.isdir(_nested):
            _append("NOTE: stale folder found: %s — delete it so old EAs don't confuse the "
                    "tester. Only files directly in %s are used." % (_nested, out_dir))
        src = source_var.get()
        _append("Generating %d EA(s) into %s ..." % (len(_selected), out_dir))
        results = batch_generate(out_dir, source=src, entries=_selected)
        ok = sum(1 for r in results if r.get("ok"))
        _append("Done: %d/%d EAs written." % (ok, len(results)))
        for r in results:
            if not r.get("ok"):
                _append("  FAILED %s: %s" % (r.get("name"), r.get("error")))
        _append("Manifest: %s" % os.path.join(out_dir, "batch_manifest.json"))
        return ok > 0
    except Exception as e:
        _append("ERROR during generate: %s" % e)
        return False

def _do_generate(source_var):
    # CHANGED 2026-06-26 — confirm before generating a large batch (the scenario
    # fan can select hundreds/thousands of variants). messagebox must run on the
    # MAIN thread, so check the selection here BEFORE handing off to the worker.
    _sel_n = len([i for i in _batch_sel_iids if i in _iid_to_entry])
    if _sel_n > 200:
        import tkinter.messagebox as _mb
        if not _mb.askyesno(
                "Generate many EAs?",
                "You're about to generate %d EAs. This will take a while and "
                "produce a large MT5 batch. Continue?" % _sel_n):
            _append("Generate cancelled (large selection).")
            return
    _run_bg(lambda: _step_generate(source_var))


def _step_compile():
    """Synchronous: returns True on success, False on error."""
    try:
        from project3_live_trading.batch_ea_tools import (
            batch_compile, get_saved_metaeditor_path,
            save_metaeditor_path, _find_metaeditor,
        )
        global _last_out_dir
        out_dir = _canonical_batch_dir()
        if not out_dir:
            _append("Compile cancelled (no MT5 data folder known).")
            return False
        _last_out_dir = out_dir
        data_dir = _derive_data_dir(out_dir)
        if not data_dir:
            _append("Compile failed: couldn't derive MT5 data dir from %s" % out_dir)
            return False
        _append("Compiling from: %s" % out_dir)

        me = get_saved_metaeditor_path() or _find_metaeditor()
        if not me:
            _append("Compile failed: metaeditor64.exe not found. Run '1b. Compile EAs' manually once to set path.")
            return False

        _append("Compiling EAs with metaeditor64 (headless) ...")
        results = batch_compile(out_dir, data_dir, metaeditor_path=me)
        ok = sum(1 for r in results if r.get("ok"))
        _append("Compiled %d/%d EA(s) to .ex5." % (ok, len(results)))
        for r in results:
            if r.get("ok"):
                _append("  OK %s (%d bytes)" % (r.get("name"), r.get("ex5_size", 0)))
        for r in results:
            if not r.get("ok"):
                _append("  FAILED %s: %s" % (r.get("name"), r.get("error")))
        if any("lock" in (r.get("error") or "").lower() or
               "WinError 32" in (r.get("error") or "") for r in results):
            _append("Some files were LOCKED. Close MT5 terminal and try again.")
            return False
        return ok == len(results) and ok > 0
    except Exception as e:
        _append("ERROR during compile: %s" % e)
        return False


def _step_build_run_files():
    """Synchronous: returns True on success, False on error."""
    try:
        from project3_live_trading.batch_ea_tools import emit_tester_inis, resolve_terminal_path
        global _last_out_dir, _last_bat_path, _last_reports_dir
        batch_dir = _canonical_batch_dir()
        if not batch_dir:
            _append("Build run files cancelled (no MT5 data folder known).")
            return False
        _last_out_dir = batch_dir
        manifest = os.path.join(batch_dir, "batch_manifest.json")
        if not os.path.isfile(manifest):
            _append("No manifest — run '1. Generate EAs' first.")
            return False
        data_dir = _derive_data_dir(batch_dir)
        if not data_dir:
            _append("Build failed: couldn't derive MT5 data dir")
            return False
        reports_dir = os.path.join(batch_dir, "reports")
        term = resolve_terminal_path()
        if not term:
            _append("Build failed: terminal64.exe not found. Run '2. Build Run Files' manually once to set path.")
            return False
        emit_tester_inis(manifest, data_dir, "Experts\\batch", reports_dir, terminal_exe=term)
        _last_bat_path = os.path.join(batch_dir, "run_all_tests.bat")
        _last_reports_dir = reports_dir
        _append("Run files written.")
        return True
    except Exception as e:
        _append("ERROR during build: %s" % e)
        return False


def _step_run_tests():
    """Synchronous: BLOCKS until bat completes. Returns True on success."""
    try:
        import subprocess, threading, time, glob as _glob, json as _json
        global _last_bat_path, _last_reports_dir
        bat = _last_bat_path
        if not bat or not os.path.isfile(bat):
            _append("Run Tests failed: no .bat (run '2. Build Run Files' first).")
            return False
        if _mt5_is_running():
            _append("MT5 is OPEN — close the terminal first.")
            return False

        _total_eas = 0
        try:
            _man = os.path.join(os.path.dirname(bat), "batch_manifest.json")
            if os.path.isfile(_man):
                with open(_man, encoding="utf-8") as _f:
                    _total_eas = len(_json.load(_f))
        except Exception:
            pass

        if _last_reports_dir and os.path.isdir(_last_reports_dir):
            _cleared = 0
            # CHANGED: June 2026 — also clear .xlsx reports (tester now writes xlsx)
            for _f in (_glob.glob(os.path.join(_last_reports_dir, "*.xlsx")) +
                       _glob.glob(os.path.join(_last_reports_dir, "*.htm")) +
                       _glob.glob(os.path.join(_last_reports_dir, "*.html"))):
                try:
                    os.remove(_f)
                    _cleared += 1
                except Exception:
                    pass

        _append("Running %d EA(s) via %s ..." % (_total_eas, os.path.basename(bat)))
        proc = subprocess.Popen(["cmd", "/c", bat], cwd=os.path.dirname(bat),
                                stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                                text=True, bufsize=1)

        _seen = {"n": 0}
        def _pump():
            for line in proc.stdout:
                line = line.rstrip()
                if not line:
                    continue
                if "=== RUNNING" in line:
                    _seen["n"] += 1
                    if _total_eas:
                        line = line.replace("=== RUNNING",
                                            "=== RUNNING [%d/%d]" % (_seen["n"], _total_eas))
                _append("  " + line)
        threading.Thread(target=_pump, daemon=True).start()

        while proc.poll() is None:
            time.sleep(5)
            # CHANGED: June 2026 — count .xlsx reports (full deal data) instead of .htm (summary only)
            # WHY: tester now writes .xlsx with complete trade details.
            done_n = (len(_glob.glob(os.path.join(_last_reports_dir, "*.xlsx"))) +
                      len(_glob.glob(os.path.join(_last_reports_dir, "*.htm"))) +
                      len(_glob.glob(os.path.join(_last_reports_dir, "*.html"))))             if (_last_reports_dir and os.path.isdir(_last_reports_dir)) else 0
            _prog = "%d/%d done" % (done_n, _total_eas) if _total_eas else "%d done" % done_n
            _append("  [heartbeat] %s" % _prog)

        _append("All tester passes finished." if proc.returncode == 0 else
                "Tester exited code %d" % proc.returncode)
        return proc.returncode == 0
    except Exception as e:
        _append("ERROR during run: %s" % e)
        return False


def _step_compare():
    """Synchronous: returns True on success."""
    try:
        from project3_live_trading.batch_compare_reports import compare_reports
        global _last_reports_dir, _last_out_dir
        reports = (_last_reports_dir if (_last_reports_dir and os.path.isdir(_last_reports_dir))
                   else None)
        if not reports:
            bdir = _last_out_dir or _canonical_batch_dir()
            if bdir:
                cand = os.path.join(bdir, "reports")
                if os.path.isdir(cand):
                    reports = cand
        if not reports:
            _append("No reports folder — run tests first.")
            return False
        manifest = ""
        bdir = _last_out_dir or _canonical_batch_dir()
        if bdir:
            m = os.path.join(bdir, "batch_manifest.json")
            if os.path.isfile(m):
                manifest = m
        _append("Comparing (%d reports)..." % len(os.listdir(reports)))
        rows = compare_reports(reports, "", manifest)
        for line in rows:
            _append(line)
        return True
    except Exception as e:
        _append("ERROR during compare: %s" % e)
        return False


def _resolve_eval_settings_from_rule(rule_dict):
    """Refiner-style per-rule resolution of (firm_id, challenge_id, account,
    risk_pct, sl_pips, pip_value, firm_label). Reads stage/firm/account/risk/SL
    from the rule itself — never hardcodes. Mirrors _update_passrate_detail in
    strategy_refiner_panel.py so the EA eval files match the refiner panel."""
    from shared.sim_detail import resolve_firm_challenge

    rd = rule_dict or {}
    rule0 = ((rd.get('rules') or [{}])[0] if rd.get('rules')
             else (rd.get('saved_rule') or {})) or {}
    rs0 = rd.get('run_settings') or {}

    def _f(*vals, default):
        for v in vals:
            if v not in (None, '', 0):
                try:
                    return float(v)
                except Exception:
                    pass
        return float(default)

    # Account: rule's own value wins; otherwise the batch panel's Account dropdown.
    try:
        _panel_acct = (int(_acct_var.get()) if (_acct_var is not None
                       and str(_acct_var.get()).strip()) else 10000)
    except Exception:
        _panel_acct = 10000
    acct = _f(rule0.get('account_size'), rs0.get('starting_capital'),
              rd.get('account_size'), default=_panel_acct)
    risk = _f(rule0.get('risk_pct'), rs0.get('risk_pct'),
              rd.get('risk_pct'), default=1.0)
    sl   = _f((rule0.get('exit_params') or {}).get('sl_pips'),
              default=150.0)
    pipv = _f(rule0.get('pip_value_per_lot'), rs0.get('pip_value_per_lot'),
              default=1.0)

    # Stage — read from the rule (informational; firm phase drives the numbers).
    stage = (rd.get('prop_firm_stage')
             or rs0.get('prop_firm_stage')
             or rule0.get('prop_firm_stage')
             or 'evaluation')

    top_firm = (rd.get('prop_firm_name') or rd.get('firm_name')
                or rs0.get('prop_firm_name') or rs0.get('firm_name'))
    top_firm_id = (rd.get('firm_id') or rs0.get('firm_id'))

    firm_id, ch_id = resolve_firm_challenge(
        rule0, int(acct),
        fallback_firm_name=top_firm,
        fallback_firm_id=top_firm_id)

    return {
        'firm_id': firm_id, 'challenge_id': ch_id,
        'account_size': int(acct), 'risk_pct': risk,
        'sl_pips': sl, 'pip_value': pipv, 'stage': stage,
        'firm_label': (top_firm or rule0.get('prop_firm_name') or firm_id or '?'),
    }


def _mt5_trades_to_sim_trades(mt5_trades, pip_value_per_lot):
    """Convert parsed MT5 trades (dollar profit + volume) into the trade-dict
    shape build_sim_detail_text/_trades_to_df expect (net_pips + exit_time).

    pips = profit / (volume * pip_value_per_lot). Uses MT5's real net profit
    (already includes commission/swap), so the eval reflects actual fills.
    Volume<=0 or missing pip value -> trade skipped (can't size it)."""
    out = []
    pv = float(pip_value_per_lot or 0)
    for t in mt5_trades or []:
        try:
            vol = float(t.get('volume') or t.get('Volume') or 0)
            profit = float(t.get('profit') or t.get('Profit') or 0)
        except Exception:
            continue
        if vol <= 0 or pv <= 0:
            continue
        net_pips = profit / (vol * pv)
        out.append({
            'entry_time': t.get('entry_time') or t.get('time'),
            'exit_time':  t.get('exit_time') or t.get('entry_time') or t.get('time'),
            'net_pips':   net_pips,
        })
    return out


def _write_sim_detail_xlsx(path, text):
    """Render build_sim_detail_text() output into an xlsx that mirrors the
    refiner 'eval & funded simulation' panel: one line per row, Consolas font,
    green for PASS rows, red for FAIL rows, grey for incomplete rows."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except Exception:
        # openpyxl missing — fall back to a .txt so the data is never lost.
        with open(path.rsplit(".", 1)[0] + ".txt", "w", encoding="utf-8") as _f:
            _f.write(text)
        return

    GREEN = PatternFill("solid", fgColor="C6EFCE")
    RED   = PatternFill("solid", fgColor="FFC7CE")
    GREY  = PatternFill("solid", fgColor="E0E0E0")
    mono  = Font(name="Consolas", size=10)
    head  = Font(name="Consolas", size=10, bold=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "eval & funded"

    for i, line in enumerate(text.split("\n"), 1):
        c = ws.cell(row=i, column=1, value=line)
        c.font = mono
        c.alignment = Alignment(horizontal="left", vertical="center")
        # header-ish lines (Firm:/EVAL/FUNDED/Per-window) get bold
        _stripped = line.lstrip()
        if (_stripped.startswith(("Firm:", "🎯", "💰", "📋"))):
            c.font = head
        # per-window outcome colouring
        if "✓ PASS" in line:
            c.fill = GREEN
        elif "✗" in line:
            c.fill = RED
        elif "… Incomplete" in line:
            c.fill = GREY

    ws.column_dimensions["A"].width = 90
    wb.save(path)


def _write_debug_dump():
    """Write MT5 + Python trades + comparison into per-EA folders AND a single
    consolidated batch_debug.json. Restores the per-EA folder layout (mt5_trades.csv,
    python_trades.csv, entry_compare.csv, summary.txt, copied report) and adds one
    aggregate file on top."""
    import json, glob, datetime, csv as _csv, shutil, re as _re
    try:
        from project3_live_trading.batch_compare_reports import (
            _load_py_trades_for, _py_rules_dir, _parse_mt5_xlsx, _parse_mt5_html,
            _find_report, _parse_mt5_html_stats, _read_mt5_entries,
            _compare, _tf_bar_minutes)
    except Exception:
        _append("Debug dump skipped (import failed).")
        return
    reports = _last_reports_dir
    bdir = _last_out_dir or _canonical_batch_dir()
    if not reports or not bdir:
        return
    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    dump = os.path.join(bdir, "debug_dump_%s" % stamp)
    os.makedirs(dump, exist_ok=True)

    _dbg_path = os.path.join(dump, "_DEBUG.txt")
    _dbg = None
    try:
        _dbg = open(_dbg_path, "w", encoding="utf-8")
    except Exception:
        pass
    def _d(msg):
        try:
            if _dbg is not None:
                _dbg.write(msg + "\n"); _dbg.flush()
        except Exception:
            pass
        _append(msg)

    man = {}
    mpath = os.path.join(bdir, "batch_manifest.json")
    if os.path.isfile(mpath):
        with open(mpath, encoding="utf-8") as f:
            for rec in json.load(f):
                man[rec.get("name")] = rec
    _d("manifest: %s (%d records)" % (mpath if os.path.isfile(mpath) else "NOT FOUND", len(man)))

    report_files = (glob.glob(os.path.join(reports, "*.xlsx.htm")) +
                    glob.glob(os.path.join(reports, "*.xlsx")) +
                    glob.glob(os.path.join(reports, "*.htm")) +
                    glob.glob(os.path.join(reports, "*.html")))
    _d("reports dir: %s" % reports)
    _d("files found: %s" % ([os.path.basename(p) for p in report_files] or "NONE"))

    # ADDED 2026-06-15 — aggregate list for the consolidated batch_debug.json
    _consolidated = {
        "generated": datetime.datetime.now().isoformat(timespec="seconds"),
        "reports_dir": reports,
        "window": {"from": "2026-01-01", "to": "2026-04-08"},
        "eas": [],
    }

    for rpt in sorted(report_files):
        base = os.path.basename(rpt)
        for _ in range(3):
            base, ext = os.path.splitext(base)
            if not ext:
                break
        ea = base

        sub = os.path.join(dump, ea)
        os.makedirs(sub, exist_ok=True)
        _d("---- %s ----" % os.path.basename(rpt))
        _d("  EA name (stripped): %s" % ea)

        # ADDED 2026-06-15 — copy the actual MT5 report (xlsx/htm) into the EA folder
        # WHY: the Excel/HTML report is the file you open to eyeball the run; it must live
        #   alongside the csvs again.
        try:
            shutil.copy(rpt, os.path.join(sub, os.path.basename(rpt)))
            _d("  copied report: %s" % os.path.basename(rpt))
        except Exception as _ce:
            _d("  report copy failed: %r" % _ce)

        mt5_trades = []
        mt5_stats = {}
        try:
            entries = _read_mt5_entries(rpt)
            _d("  _read_mt5_entries: %d entries" % len(entries))
        except Exception as e:
            _d("  _read_mt5_entries FAILED: %r" % e)
            entries = []
        try:
            if rpt.lower().endswith(".xlsx"):
                mt5_trades, mt5_stats = _parse_mt5_xlsx(rpt)
                _d("  _parse_mt5_xlsx: %d trades" % len(mt5_trades))
                if not mt5_trades and entries:
                    mt5_trades = [{"entry_time": e.strftime("%Y-%m-%d %H:%M:%S") if hasattr(e, 'strftime') else str(e)}
                                  for e in entries]
                    _d("  → FALLBACK to entry reader: %d entry-only rows" % len(mt5_trades))
            elif rpt.lower().endswith((".htm", ".html")):
                mt5_trades, mt5_stats = _parse_mt5_html(rpt)
                _d("  _parse_mt5_html: %d trades" % len(mt5_trades))
        except Exception as e:
            _d("  parse FAILED: %r" % e)

        with open(os.path.join(sub, "mt5_trades.csv"), "w", newline="", encoding="utf-8") as f:
            if mt5_trades:
                w = _csv.DictWriter(f, fieldnames=list(mt5_trades[0].keys()))
                w.writeheader()
                w.writerows(mt5_trades)

        # ADDED 2026-06-15 — robust combo matching to kill "no stored Python run (combo=?)".
        # WHY: rule/exit hashes in EA filenames are 4-char hex (e.g. 6179, 016e), not 8-char.
        #   The {8} regex missed them entirely, so hexids was always []. Using {4,8} with
        #   word-boundary lookarounds extracts 4-to-8 char hex tokens that are not part of a
        #   longer run (avoids matching '4c' in '4c6179' if ever concatenated).
        _HEX_RE = r"(?<![0-9a-f])[0-9a-f]{4,8}(?![0-9a-f])"
        rec = man.get(ea)
        if rec is None:
            _hexids = _re.findall(_HEX_RE, ea.lower())
            for k, v in man.items():
                if any(h in k.lower() or h in str(v.get("rule_combo", "")).lower() for h in _hexids):
                    rec = v
                    break
        if rec is None:
            _tf_m = _re.search(r"_(M5|M15|M30|H1|H4|D1|W1)_", ea)
            _hexids = _re.findall(_HEX_RE, ea.lower())
            # combo string the rules-dir scan will try to match against
            _combo_guess = next((h for h in _hexids), ea)
            rec = {
                "name": ea,
                "rule_combo": _combo_guess,
                "entry_tf": (_tf_m.group(1) if _tf_m else ""),
                "exit_name": "",
            }
            _d("  rec: manifest MISS — derived combo=%s tf=%s (hexids=%s)"
               % (rec["rule_combo"], rec["entry_tf"], ",".join(_hexids) or "none"))

        py, meta = (_load_py_trades_for(rec.get("rule_combo", ""), rec.get("exit_name", ""),
                                        rec.get("entry_tf", "")) if rec else (None, None))

        # ADDED 2026-06-15 — last-resort rules-dir scan when the lookup still returns nothing
        # WHY: _load_py_trades_for keys on combo; if the derived combo doesn't match its index,
        #   fall back to scanning *.json filenames for either hex id and load directly.
        if not py:
            try:
                _rdir = _py_rules_dir()
                _hexids = _re.findall(_HEX_RE, ea.lower())
                # WHY (June 2026 — TF guard): The hex-ID scan previously matched
                #      on exit hashes (e.g. 0df9) shared across rules, letting M5
                #      rule files match H4 EAs. Adding a TF check prevents this.
                # CHANGED: June 2026 — TF guard in fallback rules-dir scan
                _want_tf_scan = (rec.get("entry_tf") or "").upper().strip()
                for _jf in sorted(glob.glob(os.path.join(_rdir, "*.json"))):
                    _bn = os.path.basename(_jf).lower()
                    if any(h in _bn for h in _hexids):
                        with open(_jf, encoding="utf-8") as _fh:
                            _rd = json.load(_fh)
                        _got_tf = (_rd.get("entry_tf") or "").upper().strip()
                        if _want_tf_scan and _got_tf and _got_tf != _want_tf_scan:
                            _d("  rules-dir scan SKIP (TF mismatch): %s has %s, want %s"
                               % (os.path.basename(_jf), _got_tf, _want_tf_scan))
                            continue
                        py = _rd.get("trades") or _rd.get("py_trades") or py
                        meta = {"file": os.path.basename(_jf),
                                "py_tf": _rd.get("entry_tf"),
                                "tf_match": (_got_tf == _want_tf_scan)}
                        _d("  rules-dir scan matched: %s (%d trades)"
                           % (os.path.basename(_jf), len(py or [])))
                        break
            except Exception as _se:
                _d("  rules-dir scan failed: %r" % _se)

        if not py:
            _d("  PYTHON: no stored run resolved for combo=%s tf=%s"
               % (rec.get("rule_combo", "?"), rec.get("entry_tf", "?")))

        # ---- write python_trades.csv (origin flattening preserved) ----
        with open(os.path.join(sub, "python_trades.csv"), "w", newline="", encoding="utf-8") as f:
            if py:
                _flat_py = []
                for _t in py:
                    _row = dict(_t)
                    _dbgd = _row.pop("entry_debug", None)
                    _row.pop("entry_indicators", None)
                    if isinstance(_dbgd, dict):
                        for _feat, _info in _dbgd.items():
                            _sf = _feat.replace(' ', '_').replace('-', '_')
                            if isinstance(_info, dict):
                                _row["ind_" + _sf + "_val"] = _info.get('value')
                                _row["ind_" + _sf + "_ts"]  = _info.get('entry_row_ts')
                            else:
                                _row["ind_" + _sf] = _info
                    _flat_py.append(_row)
                _all_keys = list(dict.fromkeys(k for _r in _flat_py for k in _r.keys()))
                w = _csv.DictWriter(f, fieldnames=_all_keys, extrasaction='ignore')
                w.writeheader()
                w.writerows(_flat_py)

        # ---- summary.txt (origin content, trimmed comments) ----
        def _f(x):
            try: return float(x)
            except Exception: return None
        py_net_pips = sum((_f(t.get("net_pips")) or 0) for t in (py or [])) if py else None
        with open(os.path.join(sub, "summary.txt"), "w", encoding="utf-8") as f:
            f.write("EA: %s\ncombo: %s\nexit: %s\ntf: %s\n\n" %
                    (ea, rec.get("rule_combo", "?"), rec.get("exit_name", "?"),
                     rec.get("entry_tf", "?")))
            f.write("MT5:\n  trades: %d\n  net_profit: %s\n  profit_factor: %s\n\n" %
                    (len(mt5_trades), mt5_stats.get("net_profit", "?"),
                     mt5_stats.get("profit_factor", "?")))
            f.write("Python:\n  trades: %d\n  net_pips: %s\n  source: %s\n" %
                    (len(py or []),
                     ("%.1f" % py_net_pips) if py_net_pips is not None else "?",
                     (meta or {}).get("file", "?")))

        # ADDED 2026-06-24 — refiner-style eval, TWO files per EA:
        #   eval_windows_PY.xlsx  (Python trades — matches refiner panel exactly)
        #   eval_windows_MT5.xlsx (MT5 executed trades, same per-rule settings)
        # Per-rule resolution of stage/firm/account/risk/SL — never hardcoded.
        try:
            import json as _json_ev
            from shared.sim_detail import build_sim_detail_text

            # CHANGED 2026-06-25 — rule JSON is OPTIONAL. In backtest mode it gives
            #   per-rule stage/firm/account/risk/SL; in scenario mode (no Python
            #   run) it's absent, so we fall back to firm defaults and still build
            #   the MT5 eval from the actual MT5 trades. PY file only when py exists.
            _rd_ev = {}
            if meta and meta.get("file"):
                _rd_path = os.path.join(_py_rules_dir(), meta["file"])
                if os.path.isfile(_rd_path):
                    with open(_rd_path, encoding="utf-8") as _fh_ev:
                        _rd_ev = _json_ev.load(_fh_ev)

            _cfg = _resolve_eval_settings_from_rule(_rd_ev)
            if not _cfg.get("firm_id"):
                _cfg["firm_id"] = "leveraged"
            if not _cfg.get("challenge_id"):
                _cfg["challenge_id"] = "leveraged_standard"

            def _gen(_trades, _suffix):
                if not _trades or len(_trades) < 1:
                    _d("  eval_windows_%s skipped (no %s trades)" % (_suffix, _suffix))
                    return
                _txt, _ = build_sim_detail_text(
                    strategy_dict=_rd_ev, trades=_trades,
                    firm_id=_cfg["firm_id"], challenge_id=_cfg["challenge_id"],
                    account_size=_cfg["account_size"],
                    risk_per_trade_pct=_cfg["risk_pct"],
                    default_sl_pips=_cfg["sl_pips"],
                    pip_value_per_lot=_cfg["pip_value"],
                    symbol="XAUUSD", firm_label=_cfg["firm_label"],
                )
                # Prefix the file with the source + resolved stage/firm + EA so it's
                # unambiguous which trades and which requirements were used.
                _hdr = ("# SOURCE: %s trades  |  STAGE: %s  |  FIRM: %s / %s  |  EA: %s\n"
                        % (_suffix, _cfg["stage"], _cfg["firm_id"],
                           _cfg["challenge_id"], ea))
                _write_sim_detail_xlsx(
                    os.path.join(sub, "eval_windows_%s.xlsx" % _suffix),
                    _hdr + _txt)
                _d("  eval_windows_%s.xlsx written (stage=%s firm=%s, %d trades)"
                   % (_suffix, _cfg["stage"], _cfg["firm_id"], len(_trades)))

            # MT5 file — ALWAYS attempt (this is what scenario mode needs).
            _mt5_sim = _mt5_trades_to_sim_trades(mt5_trades, _cfg["pip_value"])
            _gen(_mt5_sim, "MT5")

            # PY file — only when a Python run exists (backtest mode).
            if py:
                _gen(py, "PY")
            else:
                _d("  eval_windows_PY skipped (no Python run — scenario mode)")
        except Exception as _ev_err:
            _d("  eval_windows generation failed: %r" % _ev_err)

        # ADDED 2026-06-15 — append this EA to the consolidated aggregate
        _consolidated["eas"].append({
            "name": ea,
            "combo": rec.get("rule_combo", "?"),
            "exit": rec.get("exit_name", "?"),
            "tf": rec.get("entry_tf", "?"),
            "mt5": {"trades": len(mt5_trades), "stats": mt5_stats},
            "python": {"trades": len(py or []), "net_pips": py_net_pips,
                       "source": (meta or {}).get("file")},
        })

    cs = os.path.join(reports, "comparison_summary.csv")
    if os.path.isfile(cs):
        shutil.copy(cs, os.path.join(dump, "comparison_summary.csv"))

    # ---- condlog + journal copy (broadened globs; origin behavior kept) ----
    _copy_data_dir = _derive_data_dir(_last_out_dir) if _last_out_dir else None
    if _copy_data_dir:
        # WHY: MT5 stores tester agent files under MetaQuotes\Tester\<hash>\ while the EA
        #      data dir (_copy_data_dir) is MetaQuotes\Terminal\<hash>\ — SIBLING directories
        #      under MetaQuotes, not nested. All prior globs searched Terminal\<hash>\Tester\...,
        #      which does not exist, so condlog and agent logs were never found.
        # CHANGED: June 2026 — add sibling Tester\<hash> path to all glob candidates
        _meta_root   = os.path.dirname(os.path.dirname(_copy_data_dir))  # ...MetaQuotes
        _data_hash   = os.path.basename(_copy_data_dir)                   # D0E8209F77C8CF37...
        _tester_root = os.path.join(_meta_root, 'Tester', _data_hash)    # Tester\<hash>
        _clogs = sorted(
            glob.glob(os.path.join(_tester_root, 'Agent-*', 'MQL5', 'Files', 'condlog_*.csv')) +
            glob.glob(os.path.join(_copy_data_dir, 'Tester', '*', 'Agent-*', 'MQL5', 'Files', 'condlog_*.csv')) +
            glob.glob(os.path.join(_copy_data_dir, 'Tester', 'Agent-*', 'MQL5', 'Files', 'condlog_*.csv')) +
            glob.glob(os.path.join(_copy_data_dir, 'MQL5', 'Files', 'condlog_*.csv')),
            key=os.path.getmtime, reverse=True)
        if _clogs:
            shutil.copy(_clogs[0], os.path.join(dump, "condlog.csv"))
            _d("Copied condlog: %s" % _clogs[0])
            # WHY: Parse condlog to find ALL-PASS bars (MT5's actual signals).
            #      Save to mt5_signal_override.json so the Python backtester can
            #      OPT IN (env MT5_SIGNAL_OVERRIDE) to supplement its signal mask.
            # CHANGED: June 2026 — condlog signal override for parity (opt-in consumer)
            try:
                _cl_bars = {}
                with open(_clogs[0], encoding='utf-8-sig') as _clf_ov:
                    for _clr_ov in _csv.DictReader(_clf_ov):
                        _bt_ov = _clr_ov.get('bar_time', '').strip()
                        _pf_ov = _clr_ov.get('pass_fail', '').strip().upper()
                        _cl_bars.setdefault(_bt_ov, []).append(_pf_ov)
                _all_pass_ov = sorted([bt for bt, pfs in _cl_bars.items()
                                       if all(p == 'PASS' for p in pfs)])
                _repo_root = os.path.dirname(os.path.dirname(os.path.dirname(
                    os.path.abspath(__file__))))
                _ov_path = os.path.join(_repo_root,
                    'project2_backtesting', 'outputs', 'mt5_signal_override.json')
                with open(_ov_path, 'w', encoding='utf-8') as _ovf:
                    json.dump({'signal_bars': _all_pass_ov,
                               'source': str(_clogs[0]),
                               'generated': str(datetime.datetime.now())}, _ovf, indent=2)
                _d("MT5 signal override: %d ALL-PASS bars -> %s" % (len(_all_pass_ov), _ov_path))
            except Exception as _ov_err:
                _d("MT5 signal override failed: %r" % _ov_err)
        else:
            _d("condlog: not found (run with DebugConditions=true first)")
        # CHANGED 2026-06-15 — select the AGENT log by CONTENT, not mtime.
        # WHY: terminal \Logs\YYYYMMDD.log is newest by mtime but has no EA prints.
        #   The tester agent log (source "Core NN") holds [DIAG]/signal=/[SKIP]. Pick the
        #   newest log that actually contains those markers; mtime is only a tiebreak.
        _jcands = sorted(
            set(
                glob.glob(os.path.join(_tester_root, 'Agent-*', 'logs', '*.log')) +
                glob.glob(os.path.join(_tester_root, 'Agent-*', 'Logs', '*.log')) +
                glob.glob(os.path.join(_copy_data_dir, 'Tester', '*', 'Agent-*', 'logs', '*.log')) +
                glob.glob(os.path.join(_copy_data_dir, 'Tester', '*', 'Agent-*', 'Logs', '*.log')) +
                glob.glob(os.path.join(_copy_data_dir, 'Tester', 'Agent-*', 'logs', '*.log')) +
                glob.glob(os.path.join(_copy_data_dir, 'Tester', 'Agent-*', 'Logs', '*.log')) +
                glob.glob(os.path.join(_copy_data_dir, 'Tester', '*', 'Agent-*', '*.log')) +
                glob.glob(os.path.join(_copy_data_dir, 'Logs', '*.log'))
            ),
            key=os.path.getmtime, reverse=True
        )

        def _has_ea_prints(path):
            # Read a chunk and check for EA diagnostic markers. MT5 logs are UTF-16LE.
            try:
                with open(path, 'rb') as _fh:
                    raw = _fh.read(2_000_000)  # 2 MB sample is plenty
                for enc in ('utf-16', 'utf-8', 'cp1252'):
                    try:
                        txt = raw.decode(enc, errors='ignore')
                        break
                    except Exception:
                        txt = ''
                return any(m in txt for m in ('[DIAG]', 'signal=', '[SKIP]', '[LOTS]'))
            except Exception:
                return False

        _agent_log = next((p for p in _jcands if _has_ea_prints(p)), None)
        _picked = _agent_log or (_jcands[0] if _jcands else None)
        if _picked:
            shutil.copy(_picked, os.path.join(dump, "mt5_journal.log"))
            if _agent_log:
                _d("Copied AGENT journal (has EA prints): %s" % _picked)
            else:
                _d("Copied journal (NO EA prints found — likely terminal log; "
                   "run with DebugConditions=true and check Tester agent ran): %s" % _picked)
        else:
            _d("journal: no *.log found in Tester agent folders or \\Logs")

    # ADDED 2026-06-15 — write the consolidated aggregate alongside the folders
    try:
        with open(os.path.join(dump, "batch_debug.json"), "w", encoding="utf-8") as f:
            json.dump(_consolidated, f, indent=1, default=str)
        _d("Consolidated dump: batch_debug.json (%d EAs) — per-EA folders ALSO written"
           % len(_consolidated["eas"]))
    except Exception as _je:
        _d("consolidated write failed: %r" % _je)

    # ---- decode journal once (shared by PARITY_BUNDLE and PARITY_REPORT) ----
    _jraw, _jtxt = b"", ""
    _jpath = os.path.join(dump, "mt5_journal.log")
    if os.path.isfile(_jpath):
        try:
            with open(_jpath, "rb") as _jf:
                _jraw = _jf.read()
            for _enc in ("utf-16", "utf-8", "cp1252"):
                try:
                    _jtxt = _jraw.decode(_enc, errors="ignore")
                    if "ENTRY-EVAL" in _jtxt or "[DIAG]" in _jtxt or "[ATR-EXIT]" in _jtxt:
                        break
                except Exception:
                    _jtxt = ""
        except Exception:
            _jtxt = ""
    _jlines = _jtxt.splitlines()

    def _slice_for(ea_name, tags, max_lines=80):
        out = []
        for ln in _jlines:
            if ea_name in ln and any(t in ln for t in tags):
                out.append(ln.rstrip())
                if len(out) >= max_lines:
                    out.append("  ... (truncated at %d)" % max_lines)
                    break
        return out

    def _deep_get(o, key):
        # WHY: entry_bar_offset and run_max_spread_pips may live at the top level
        #   or inside a nested dict depending on when the rule JSON was written.
        # CHANGED: 2026-06-17 — recursive key lookup for bundle correctness
        if isinstance(o, dict):
            if key in o:
                return o[key]
            for v in o.values():
                r = _deep_get(v, key)
                if r is not None:
                    return r
        return None

    # ADDED 2026-06-17 — single self-contained PARITY_BUNDLE.txt.
    # WHY: one upload instead of many scattered files. Inlines comparison_summary,
    #   per-EA summary + MT5 + Python trades + exit params, and the per-EA AGENT-LOG
    #   SLICE (decoded from the UTF-16 journal already copied above). Purely additive;
    #   per-EA folders and batch_debug.json are untouched.
    try:
        _bundle_path = os.path.join(dump, "PARITY_BUNDLE.txt")
        with open(_bundle_path, "w", encoding="utf-8") as _bf:
            _bf.write("PARITY BUNDLE  generated %s\n" % _consolidated["generated"])
            _bf.write("window %s..%s   reports_dir %s\n" %
                      (_consolidated["window"]["from"], _consolidated["window"]["to"], reports))
            _bf.write("journal lines decoded: %d\n" % len(_jlines))
            # Probe: is the SESSIONGAP gap-fill code present in the LOADED backtester?
            # WHY: if the fix isn't active entries stay at 04:00 — surface it so we
            #   never again wonder whether the code is loaded vs the data is missing.
            try:
                import inspect as _insp
                from project2_backtesting import strategy_backtester as _sbt
                _gap_present = "gap_fill_parity" in _insp.getsource(_sbt)
                _bf.write("gap-fill code in loaded backtester: %s\n"
                          % ("PRESENT" if _gap_present else "ABSENT (entries will stay at boundary)"))
            except Exception as _ge:
                _bf.write("gap-fill code probe failed: %r\n" % _ge)
            _bf.write("=" * 78 + "\n\n")

            # ---- comparison_summary.csv verbatim ----
            _csv_path = os.path.join(dump, "comparison_summary.csv")
            _bf.write("### COMPARISON_SUMMARY.CSV\n")
            if os.path.isfile(_csv_path):
                with open(_csv_path, encoding="utf-8") as _cf:
                    _bf.write(_cf.read().strip() + "\n")
            else:
                _bf.write("(not found)\n")
            _bf.write("\n" + "=" * 78 + "\n\n")

            # ---- per-EA section ----
            for _e in _consolidated["eas"]:
                _ea = _e["name"]
                _sub = os.path.join(dump, _ea)
                _bf.write("### EA: %s\n" % _ea)
                _bf.write("combo=%s exit=%s tf=%s\n" % (_e["combo"], _e["exit"], _e["tf"]))
                _bf.write("MT5 trades=%s  net_profit=%s  pf=%s\n" % (
                    _e["mt5"]["trades"],
                    _e["mt5"]["stats"].get("net_profit", "?"),
                    _e["mt5"]["stats"].get("profit_factor", "?")))
                _bf.write("PY  trades=%s  net_pips=%s  source=%s\n\n" % (
                    _e["python"]["trades"], _e["python"]["net_pips"], _e["python"]["source"]))

                _src = _e["python"]["source"]
                if _src:
                    try:
                        _rd = _py_rules_dir()
                        _sp = os.path.join(_rd, _src)
                        if os.path.isfile(_sp):
                            with open(_sp, encoding="utf-8") as _sf:
                                _sj = json.load(_sf)
                            _bf.write("PY rule: entry_bar_offset=%s spread_pips=%s "
                                      "run_max_spread_pips=%s exit_class=%s exit_params=%s\n\n" % (
                                _deep_get(_sj, "entry_bar_offset"),
                                _sj.get("spread_pips"),
                                _deep_get(_sj, "run_max_spread_pips") or "ABSENT",
                                _sj.get("exit_class"), json.dumps(_sj.get("exit_params", {}))))
                    except Exception as _xe:
                        _bf.write("PY rule params: (read failed: %r)\n\n" % _xe)

                _bf.write("-- MT5 trades --\n")
                _mp = os.path.join(_sub, "mt5_trades.csv")
                if os.path.isfile(_mp):
                    with open(_mp, encoding="utf-8") as _mf:
                        _bf.write(_mf.read().strip() + "\n")
                _bf.write("\n-- PYTHON trades --\n")
                _pp = os.path.join(_sub, "python_trades.csv")
                if os.path.isfile(_pp):
                    with open(_pp, encoding="utf-8") as _pf:
                        _bf.write(_pf.read().strip() + "\n")

                _bf.write("\n-- AGENT LOG [ATR-EXIT] / [SKIP] / signal= for this EA --\n")
                _sl = _slice_for(_ea, ("[ATR-EXIT]", "[SKIP]", "signal=true", "OnInit"))
                if _sl:
                    _bf.write("\n".join(_sl) + "\n")
                else:
                    _bf.write("(no matching agent-log lines — check DebugConditions / EA name)\n")

                # ---- SESSIONGAP DIAG (auto) ----
                # WHY: surface the session-open fill divergence every run. MT5 fills the reopen
                #   bar at first-tick time (e.g. 01:05); Python can only fill at an existing H4
                #   bar timestamp (all_times[_eb]) so it maps to the next boundary (04:00) or
                #   drops the trade. This block lists per gap fill: MT5 time, what Python did,
                #   and the FirstNewBar context — the exact inputs the SESSIONGAP fix needs.
                import csv as _csv_sg
                _tf_e = (_e.get("tf") or "H4").upper()
                _TFm_sg = {"D1":1440,"H4":240,"H1":60,"M30":30,"M15":15,"M5":5,"W1":10080}.get(_tf_e, 5)
                def _onb_sg(_ts):
                    _m_sg = _re.search(r'\s(\d{2}):(\d{2}):', _ts or "")
                    if not _m_sg: return True
                    return ((int(_m_sg.group(1))*60 + int(_m_sg.group(2))) % _TFm_sg) == 0
                _mt5_rows_sg, _py_rows_sg = [], []
                try:
                    with open(os.path.join(_sub, "mt5_trades.csv"), encoding="utf-8") as _f_sg:
                        _mt5_rows_sg = list(_csv_sg.DictReader(_f_sg))
                except Exception:
                    pass
                try:
                    with open(os.path.join(_sub, "python_trades.csv"), encoding="utf-8") as _f_sg:
                        _py_rows_sg = list(_csv_sg.DictReader(_f_sg))
                except Exception:
                    pass
                _py_ets_sg = {str(_r.get("entry_time","")).strip().replace("-",".") for _r in _py_rows_sg}
                _py_days_sg = {}
                for _r_sg in _py_rows_sg:
                    _t_sg = str(_r_sg.get("entry_time","")).strip().replace("-",".")
                    _py_days_sg.setdefault(_t_sg[:10], []).append(_t_sg[11:])
                # FirstNewBar lines for this EA from current-run journal
                _fnb_sg = {}
                for _ln_sg in [l for l in _jlines if _ea in l and "FirstNewBar" in l]:
                    _p_sg = _ln_sg.split("\t")
                    _tsf_sg = _p_sg[3] if len(_p_sg) > 3 else ""
                    _mm_sg = _re.match(r'\s*(\d{4}\.\d{2}\.\d{2}\s+\d{2}:\d{2}:\d{2})', _tsf_sg)
                    _sv_sg = _re.search(r'server=(\S+ \S+)\s+server_prev=(\S+ \S+)', _ln_sg)
                    if _mm_sg:
                        _fnb_sg[_mm_sg.group(1).strip()] = (
                            (_sv_sg.group(1), _sv_sg.group(2)) if _sv_sg else ("?","?"))
                _gap_sg = [str(_r.get("entry_time","")).strip() for _r in _mt5_rows_sg
                           if not _onb_sg(str(_r.get("entry_time","")))]
                _bf.write("\n-- SESSIONGAP DIAG (tf=%s, off-boundary MT5 fills=%d) --\n"
                          % (_tf_e, len(_gap_sg)))
                if not _gap_sg:
                    _bf.write("  (no off-boundary MT5 fills — not a session-gap EA)\n")
                for _g_sg in _gap_sg:
                    _day_sg = _g_sg[:10]
                    _matched_sg = _g_sg in _py_ets_sg
                    _sameday_sg = _py_days_sg.get(_day_sg, [])
                    if _matched_sg:
                        _verdict_sg = "PY MATCHED (same time)"
                    elif _sameday_sg:
                        _verdict_sg = "PY shifted -> %s" % ",".join(_sameday_sg)
                    else:
                        _verdict_sg = "PY DROPPED (no same-day entry)"
                    _ctx_sg = _fnb_sg.get(_g_sg)
                    _ctxs_sg = (" | FirstNewBar server=%s prev=%s" % _ctx_sg) if _ctx_sg else ""
                    _bf.write("  MT5 %s  ->  %s%s\n" % (_g_sg, _verdict_sg, _ctxs_sg))
                if _gap_sg:
                    _ym_sg = _gap_sg[0][:7].replace(".", "_")
                    try:
                        _dd_sg = _derive_data_dir(_last_out_dir) if _last_out_dir else None
                        _m1_hit_sg = ""
                        import glob as _glob_sg
                        _names_sg = ("xauusd_M1.csv", "XAUUSD_M1.csv", "M1.csv")
                        _found_sg = None
                        # 1. Search MT5 data_dir tree (original probe)
                        if _dd_sg:
                            for _nm_sg in _names_sg:
                                _hits_sg = (_glob_sg.glob(os.path.join(_dd_sg, _nm_sg)) +
                                            _glob_sg.glob(os.path.join(_dd_sg, "**", _nm_sg), recursive=True))
                                if _hits_sg:
                                    _found_sg = _hits_sg[0]
                                    break
                        # 2. Repo-anchored fallback — mirrors strategy_backtester.py 483-505.
                        #    Uses __file__ to find <repo>/data/ regardless of MT5 data_dir.
                        #    __file__ is project3_live_trading/panels/batch_eas_panel.py, so
                        #    THREE dirname() calls reach the repo root -> data/.
                        if not _found_sg:
                            try:
                                _repo_data_sg = os.path.join(
                                    os.path.dirname(os.path.dirname(os.path.dirname(
                                        os.path.abspath(__file__)))), 'data')
                                # Top-level data/<name>
                                for _nm_sg in _names_sg:
                                    _rp_sg = os.path.join(_repo_data_sg, _nm_sg)
                                    if os.path.exists(_rp_sg):
                                        _found_sg = _rp_sg
                                        break
                                # data/<d1>/<name> and data/<d1>/<d2>/<name> — mirrors the
                                # backtester's two-level scan (strategy_backtester.py 493-503),
                                # i.e. the standard data/sources/<datasource>/XAUUSD_M1.csv layout.
                                if not _found_sg and os.path.isdir(_repo_data_sg):
                                    for _d1_sg in os.scandir(_repo_data_sg):
                                        if not _d1_sg.is_dir():
                                            continue
                                        for _nm_sg in _names_sg:
                                            _rp_sg = os.path.join(_d1_sg.path, _nm_sg)
                                            if os.path.exists(_rp_sg):
                                                _found_sg = _rp_sg
                                                break
                                        if _found_sg:
                                            break
                                        for _d2_sg in os.scandir(_d1_sg.path):
                                            if not _d2_sg.is_dir():
                                                continue
                                            for _nm_sg in _names_sg:
                                                _rp_sg = os.path.join(_d2_sg.path, _nm_sg)
                                                if os.path.exists(_rp_sg):
                                                    _found_sg = _rp_sg
                                                    break
                                            if _found_sg:
                                                break
                                        if _found_sg:
                                            break
                            except Exception:
                                pass
                        if _found_sg:
                            # Detect LFS stub (3-line pointer file, < 200 bytes)
                            _sz_sg = os.path.getsize(_found_sg)
                            if _sz_sg < 200:
                                _m1_hit_sg = "%s (LFS STUB - run: git lfs pull)" % _found_sg
                            else:
                                _m1_hit_sg = _found_sg
                        else:
                            _m1_hit_sg = "NONE"
                        _bf.write("  M1 data file: %s  (backtester needs this to fill at session open)\n"
                                  % (_m1_hit_sg or "data_dir unknown"))
                    except Exception:
                        pass
                # ---- regression fingerprint (auto) ----
                # WHY: the gap fix must ONLY change gap-bar entries. REGULAR (on-boundary,
                #   non-gap-day) Python entries must be byte-identical run-to-run. Print a
                #   compact fingerprint of the REGULAR entries so a regression (gap logic
                #   leaking onto normal bars) is visible without a separate script.
                _gap_days_sg = {g[:10] for g in _gap_sg}
                _reg_entries_sg, _gap_entries_sg = [], []
                for _r_fp in _py_rows_sg:
                    _t_fp = str(_r_fp.get("entry_time", "")).strip()
                    _tn_fp = _t_fp.replace("-", ".")
                    _onbound_fp = _onb_sg(_t_fp)
                    _isgapday_fp = _tn_fp[:10] in _gap_days_sg
                    if _onbound_fp and not _isgapday_fp:
                        _reg_entries_sg.append(_tn_fp)
                    else:
                        _gap_entries_sg.append(_tn_fp)
                import hashlib as _hl_sg
                _reg_fp_sg = _hl_sg.md5(("|".join(sorted(_reg_entries_sg))).encode()).hexdigest()[:10]
                _bf.write("  -- entry split: %d regular, %d gap/off-boundary\n"
                          % (len(_reg_entries_sg), len(_gap_entries_sg)))
                _bf.write("  -- REGULAR fingerprint: %s  (must stay constant across gap-fix runs)\n"
                          % _reg_fp_sg)
                _bf.write("  -- REGULAR entries: %s\n"
                          % (", ".join(_reg_entries_sg[:12]) + (" ..." if len(_reg_entries_sg) > 12 else "")))
                # ---- end regression fingerprint ----
                # ---- end SESSIONGAP DIAG ----

                # ---- SIGNAL DIAG: Python indicator values at MT5 entry bars ----
                # WHY: When Python has fewer trades than MT5, the signal mask didn't
                #      fire on some bars. This shows Python's indicator values at each
                #      MT5 entry bar alongside MT5's condlog values, pinpointing which
                #      indicator diverges and by how much.
                # CHANGED: June 2026 — signal parity diagnostic
                try:
                    _sig_debug = _sj.get("signal_debug") if '_sj' in dir() else None
                    if _sig_debug and _mt5_rows_sg:
                        _py_sig_bars = set(_sig_debug.get("signal_bars", []))
                        _sig_ind_vals = _sig_debug.get("indicator_values", {})
                        # Load condlog for MT5 indicator values
                        _cl_path = os.path.join(dump, "condlog.csv")
                        _cl_by_bar = {}
                        if os.path.isfile(_cl_path):
                            import csv as _csv_cl
                            with open(_cl_path, encoding="utf-8-sig") as _clf:
                                for _clr in _csv_cl.DictReader(_clf):
                                    _bt = _clr.get("bar_time", "").strip()
                                    _cl_by_bar.setdefault(_bt, []).append(_clr)
                        _bf.write("\n-- SIGNAL DIAG (Python signals=%d, MT5 entries=%d) --\n"
                                  % (_sig_debug.get("signal_count", 0), len(_mt5_rows_sg)))
                        _bf.write("  %-22s %-8s  %-20s %10s %10s %10s  %s\n" % (
                            "MT5 entry", "PY sig?", "feature", "PY val", "MT5 val", "delta", "cond"))
                        for _mr in _mt5_rows_sg:
                            _mt = str(_mr.get("entry_time", "")).strip()
                            # Normalize to match signal_bars format
                            _mt_norm = _mt.replace(".", "-").replace(" ", " ")[:19]
                            _has_py = any(_mt_norm[:16] in sb[:16] for sb in _py_sig_bars)
                            _marker = "✓" if _has_py else "✗ MISS"
                            # Get Python indicator values at this bar
                            _py_vals = None
                            for _sk in _sig_ind_vals:
                                if _mt_norm[:16] in _sk[:16]:
                                    _py_vals = _sig_ind_vals[_sk]
                                    break
                            # Get MT5 condlog values at this bar
                            _mt_cl = _cl_by_bar.get(_mt, [])
                            if not _has_py:
                                _bf.write("  %-22s %-8s" % (_mt, _marker))
                                if _mt_cl:
                                    for _clr in _mt_cl:
                                        _feat = _clr.get("feature", "?")
                                        _mt5v = float(_clr.get("value", 0))
                                        _thresh = float(_clr.get("threshold", 0))
                                        _pf = _clr.get("pass_fail", "?")
                                        # Find Python's value for this feature
                                        _pyv = "?"
                                        if _py_vals:
                                            _pyv = _py_vals.get(_feat, "?")
                                        _delta = ""
                                        if isinstance(_pyv, (int, float)):
                                            _delta = "%+.4f" % (_pyv - _mt5v)
                                            _pyv = "%.4f" % _pyv
                                        _bf.write("  %-20s %10s %10.4f %10s  mt5=%s\n" % (
                                            _feat, _pyv, _mt5v, _delta, _pf))
                                        _bf.write("  %-22s %-8s" % ("", ""))
                                    _bf.write("\n")
                                else:
                                    _bf.write("  (no condlog for this bar)\n")
                            else:
                                _bf.write("  %-22s %-8s  (matched)\n" % (_mt, _marker))
                except Exception as _sde:
                    _bf.write("\n-- SIGNAL DIAG: failed (%r) --\n" % _sde)
                # ---- end SIGNAL DIAG ----

                _bf.write("\n" + "=" * 78 + "\n\n")

        _d("PARITY BUNDLE: %s" % _bundle_path)
    except Exception as _be:
        _d("PARITY_BUNDLE write failed: %r" % _be)

    # ADDED 2026-06-17 — PARITY_REPORT.txt: compact classified report, scales to 100+ EAs.
    # WHY: at 100 EAs PARITY_BUNDLE (full trade tables + 80-line log slices) is unreadable.
    #   PARITY_REPORT gives the same systemic-cause triage in ~3 lines/EA with no inline data.
    #   Tags derive from comparison_summary.csv (already has bar-size fix applied) + agent-log
    #   scan. No trade tables, no log lines — just TALLY + per-EA tag row + INDEX.
    #   PARITY_BUNDLE and per-EA folders are unchanged; this is a third additive output.
    try:
        import csv as _csv2
        from datetime import timedelta as _td2, datetime as _dt2
        from collections import Counter as _Counter

        _TF_MIN2 = {"D1": 1440, "H4": 240, "H1": 60, "M30": 30, "M15": 15, "M5": 5, "W1": 10080}

        # Parse comparison_summary.csv — has _compare() outputs with bar-size fix applied.
        _cmp_by_ea = {}
        _cs_path2 = os.path.join(dump, "comparison_summary.csv")
        if os.path.isfile(_cs_path2):
            with open(_cs_path2, encoding="utf-8") as _cf2:
                for _crow2 in _csv2.DictReader(_cf2):
                    _cmp_by_ea[_crow2.get("name", "")] = _crow2

        _rpt_rows = []
        for _e in _consolidated["eas"]:
            _ea = _e["name"]
            _sub = os.path.join(dump, _ea)
            _tf  = _e["tf"]
            try:
                _crow  = _cmp_by_ea.get(_ea, {})
                _mt5_n = int(_crow.get("mt5",         0) or 0)
                _py_n  = int(_crow.get("py",          0) or 0)
                _exct  = int(_crow.get("exact",       0) or 0)
                _inrng = int(_crow.get("py_in_range", 0) or 0)
                _late  = int(_crow.get("one_bar_late",  0) or 0)
                _after = int(_crow.get("py_after_end",  0) or 0)

                # Load rule JSON for spread metadata
                _sj3 = {}
                _src3 = _e["python"]["source"]
                if _src3:
                    _sp3 = os.path.join(_py_rules_dir(), _src3)
                    if os.path.isfile(_sp3):
                        try:
                            with open(_sp3, encoding="utf-8") as _sf3:
                                _sj3 = json.load(_sf3)
                        except Exception:
                            pass
                _py_max_spread3 = _deep_get(_sj3, "run_max_spread_pips") or 0

                # Load per-EA CSVs for EXIT_INTRABAR_SL and first-div walk
                _mt5r3, _pyr3 = [], []
                try:
                    with open(os.path.join(_sub, "mt5_trades.csv"), encoding="utf-8") as _mf3:
                        _mt5r3 = list(_csv2.DictReader(_mf3))
                except Exception:
                    pass
                try:
                    with open(os.path.join(_sub, "python_trades.csv"), encoding="utf-8") as _pf3:
                        _pyr3 = list(_csv2.DictReader(_pf3))
                except Exception:
                    pass

                # Change A — restrict to CURRENT run's lines only.
                # WHY: the agent journal accumulates across reruns; a prior run with
                #   indFail=true would mis-tag UNMAPPED even after the indicator is fixed.
                #   Walk backward from the last line while time is monotonic; a backward
                #   clock jump marks a previous-run boundary.
                _ea_all = [ln for ln in _jlines if _ea in ln]
                def _hms3(ln):
                    _p = ln.split("\t")
                    return _p[2].strip() if len(_p) > 2 else ""
                _ea_jl = _ea_all
                if _ea_all:
                    _last_t = _hms3(_ea_all[-1])
                    _cut = 0
                    for _k in range(len(_ea_all) - 1, -1, -1):
                        _t = _hms3(_ea_all[_k])
                        if _t and _last_t and _t > _last_t:
                            _cut = _k + 1
                            break
                        _last_t = _t
                    _ea_jl = _ea_all[_cut:]

                # --- classify (all applicable tags, not just primary) ---
                _tags = []
                if (_mt5_n > 0 and _exct == _mt5_n and _py_n == _mt5_n
                        and not _crow.get("error")):
                    _tags.append("CLEAN")
                else:
                    if _after > 0:
                        _tags.append("PY_AFTER_END")
                    if _late > 0:
                        # SESSIONGAP signal = MT5 entry NOT on an entry-TF bar boundary.
                        # WHY: XAUUSD reopens ~01:05 server time after the overnight/weekend
                        #   break; MT5 fills at that first-tick time while Python snaps to the
                        #   next clean H4 boundary (04:00). Off-boundary entry timestamps are
                        #   session-open fills by definition — deterministic, no log needed.
                        #   The old FirstNewBar-marker test missed most of them: the EA logs
                        #   FirstNewBar only on the backtest-start bar, not every daily reopen.
                        # 1BAR = remaining one-bar-late offset on clean boundary bars. Fix = offset=1.
                        # Both can coexist on the same EA.
                        _TF_MIN_SG = {"D1": 1440, "H4": 240, "H1": 60, "M30": 30,
                                      "M15": 15, "M5": 5, "W1": 10080}
                        _tfm_sg = _TF_MIN_SG.get((_tf or "H4").upper(), 5)
                        def _on_boundary(_ts):
                            _mm_sg = _re.search(r'\s(\d{2}):(\d{2}):', _ts)
                            if not _mm_sg:
                                return True   # can't parse → don't flag as gap
                            _mins_sg = int(_mm_sg.group(1)) * 60 + int(_mm_sg.group(2))
                            return (_mins_sg % _tfm_sg) == 0
                        _mt5_ets = {str(_mr3.get("entry_time", "")).strip() for _mr3 in _mt5r3}
                        _py_ets_norm = {
                            str(_pr3.get("entry_time", "")).strip().replace("-", ".")
                            for _pr3 in _pyr3}
                        # gap fills = off-boundary MT5 entries Python did NOT match
                        _gap_unmatched = [t for t in _mt5_ets
                                          if not _on_boundary(t) and t not in _py_ets_norm]
                        if _gap_unmatched:
                            _tags.append("ENTRY_OFFSET_SESSIONGAP")
                        # 1BAR = late offset not explained by gap fills
                        if _late > len(_gap_unmatched):
                            _tags.append("ENTRY_OFFSET_1BAR")
                        # fallback: if late>0 but neither fired (no CSVs / parse miss)
                        if ("ENTRY_OFFSET_SESSIONGAP" not in _tags
                                and "ENTRY_OFFSET_1BAR" not in _tags):
                            _tags.append("ENTRY_OFFSET_1BAR")
                    # Change C — require MT5 actually skipped a bar for spread before tagging.
                    # WHY: py_max_spread=0 alone fires on every pre-fix run; adding the
                    #   spread_too_wide agent-log confirmation makes the tag meaningful.
                    if (not _py_max_spread3
                            and any("spread_too_wide" in _ln3 for _ln3 in _ea_jl)):
                        _tags.append("SPREAD_NOT_ENFORCED")
                    # Change B — gate UNMAPPED_INDICATOR: only fire when MT5 took 0 trades AND
                    # a feature is stuck at 0.0000 on a line flagged indFail=true. Previous
                    # code matched any "indFail" substring (fires on "indFail=false" too) and
                    # ran on all lines including stale prior-run ones.
                    if _mt5_n == 0:
                        _unmapped_feat = None
                        for _ln3 in _ea_jl:
                            if "indFail=true" not in _ln3:
                                continue
                            _zeros = _re.findall(r'(\w+)=0\.0000\b', _ln3)
                            if _zeros:
                                _unmapped_feat = _zeros[0]
                                break
                        if _unmapped_feat is not None:
                            _tags.append("UNMAPPED_INDICATOR(%s)" % _unmapped_feat)
                        elif any("indicator_not_ready" in _ln3 for _ln3 in _ea_jl):
                            _tags.append("UNMAPPED_INDICATOR")
                    # ONINIT_CRASH
                    if any("OnInit" in _ln3 and
                           any(_kw3 in _ln3.lower() for _kw3 in ("critical", "fail", "error"))
                           for _ln3 in _ea_jl):
                        _tags.append("ONINIT_CRASH")
                    elif _mt5_n == 0 and _crow.get("error"):
                        _tags.append("ONINIT_CRASH")
                    # EXIT_INTRABAR_SL — matched entry: MT5 closed at loss, PY took profit
                    _mt5_eset = {str(_mr3.get("entry_time", "")).strip() for _mr3 in _mt5r3}
                    for _pr3 in _pyr3:
                        _pet3 = str(_pr3.get("entry_time", "")).strip()
                        _pex3 = str(_pr3.get("exit_reason", "")).upper()
                        _pvm3 = str(_pr3.get("exit_via_m1", "")).lower()
                        if (_pet3 in _mt5_eset and ("TP" in _pex3 or "TAKE_PROFIT" in _pex3)
                                and _pvm3 != "true"):
                            for _mr3 in _mt5r3:
                                if str(_mr3.get("entry_time", "")).strip() == _pet3:
                                    try:
                                        _m5p3 = float(str(
                                            _mr3.get("profit", _mr3.get("Profit", "0"))
                                        ).replace(",", "").strip() or "0")
                                        if _m5p3 < 0:
                                            _tags.append("EXIT_INTRABAR_SL")
                                    except Exception:
                                        pass
                                    break
                        if "EXIT_INTRABAR_SL" in _tags:
                            break
                    # LONG_HOLD_BLOCK
                    if any("position_already_open" in _ln3 for _ln3 in _ea_jl):
                        _tags.append("LONG_HOLD_BLOCK")
                    if not _tags:
                        _tags.append("UNCLASSIFIED")

                # --- first divergence (walk entry times in order, TF-tolerant) ---
                _fdiv = ""
                try:
                    def _pt3(s):
                        try:
                            return _dt2.fromisoformat(
                                str(s).split(".")[0].replace(" ", "T"))
                        except Exception:
                            return None
                    _bmin3 = _TF_MIN2.get(_tf.upper(), 5)
                    _tol3  = _td2(minutes=_bmin3)
                    _ms3   = sorted(filter(None, (_pt3(_r.get("entry_time")) for _r in _mt5r3)))
                    _ps3   = sorted(filter(None, (_pt3(_r.get("entry_time")) for _r in _pyr3)))
                    if not _ms3 and not _ps3:
                        _fdiv = "no trades in either engine"
                    elif not _ms3:
                        _fdiv = "%s PY-only entry" % _ps3[0].strftime("%Y-%m-%d %H:%M")
                    elif not _ps3:
                        _fdiv = "%s MT5-only entry" % _ms3[0].strftime("%Y-%m-%d %H:%M")
                    else:
                        _mset3 = set(_ms3)
                        _pset3 = set(_ps3)
                        for _t3 in sorted(_mset3 | _pset3):
                            _near_p3 = any(
                                abs((_t3 - _p3).total_seconds()) <= _tol3.total_seconds()
                                for _p3 in _pset3)
                            _near_m3 = any(
                                abs((_t3 - _m3).total_seconds()) <= _tol3.total_seconds()
                                for _m3 in _mset3)
                            if _t3 in _mset3 and not _near_p3:
                                _fdiv = "%s MT5-only entry" % _t3.strftime("%Y-%m-%d %H:%M")
                                break
                            if _t3 in _pset3 and not _near_m3:
                                _why3 = ""
                                for _ln3 in _ea_jl:
                                    if _t3.strftime("%Y-%m-%d") in _ln3 and "[SKIP]" in _ln3:
                                        _rm3 = _re.search(r'\[SKIP\]\s+(\S+)', _ln3)
                                        if _rm3:
                                            _why3 = _rm3.group(1)
                                        break
                                _fdiv = "%s PY-only%s" % (
                                    _t3.strftime("%Y-%m-%d %H:%M"),
                                    " MT5 skip %s" % _why3 if _why3 else " entry")
                                break
                        if not _fdiv and "EXIT_INTRABAR_SL" in _tags and _pyr3:
                            _first_pet = str(_pyr3[0].get("entry_time", ""))[:16]
                            _fdiv = "%s exit mismatch (MT5 SL, PY TP)" % _first_pet
                except Exception:
                    pass

                _rpt_rows.append({
                    "ea": _ea, "mt5": _mt5_n, "py": _py_n, "inrange": _inrng,
                    "exact": _exct, "tags": _tags, "fdiv": _fdiv, "sub": _sub,
                })
            except Exception as _re3:
                _rpt_rows.append({
                    "ea": _ea, "mt5": 0, "py": 0, "inrange": 0, "exact": 0,
                    "tags": ["UNCLASSIFIED"],
                    "fdiv": "classify-err: %r" % _re3, "sub": "",
                })

        # Build TALLY — systemic CAUSES only. PY_AFTER_END is a window artifact
        # (Python covers more data than the MT5 test window), not a divergence to fix.
        # CLEAN/UNCLASSIFIED are states, not causes. Exclude all three from the tally
        # so the ranking reflects actionable fixes. They still appear on per-EA rows.
        _TALLY_EXCLUDE = {"PY_AFTER_END", "CLEAN", "UNCLASSIFIED"}
        _tag_ctr = _Counter()
        _in_window_clean = 0        # EAs whose only non-CLEAN tag is PY_AFTER_END
        for _rr in _rpt_rows:
            _causes = [t for t in _rr["tags"] if t.split("(")[0] not in _TALLY_EXCLUDE]
            for _tg in _causes:
                _tag_ctr[_tg] += 1
            if not _causes and ("PY_AFTER_END" in _rr["tags"] or "CLEAN" in _rr["tags"]):
                _in_window_clean += 1

        _report_path = os.path.join(dump, "PARITY_REPORT.txt")
        with open(_report_path, "w", encoding="utf-8") as _rf:
            _rf.write("PARITY REPORT  generated %s\n" % _consolidated["generated"])
            _rf.write("window %s..%s  reports_dir %s\n\n" % (
                _consolidated["window"]["from"], _consolidated["window"]["to"], reports))

            # Section 1 — TALLY: systemic-fix driver (fix cause X → N EAs improve at once).
            # PY_AFTER_END/CLEAN excluded (window artifact / state, not a cause to fix).
            _n_eas = len(_rpt_rows)
            _rf.write("=== TALLY (%d EA%s — systemic causes only) ===\n" % (
                _n_eas, "s" if _n_eas != 1 else ""))
            for _tg, _cnt in sorted(_tag_ctr.items(), key=lambda _x: -_x[1]):
                _rf.write("%-36s %d\n" % (_tg, _cnt))
            _rf.write("%-36s %d\n" % ("(in-window clean / after-end only)", _in_window_clean))
            _rf.write("\n")

            # Section 2 — PER-EA ROWS (one line each; causes lead, PY_AFTER_END/CLEAN trail)
            _rf.write("=== PER-EA ROWS ===\n")
            _rf.write("%-52s %4s %4s %7s %5s  %-52s  %s\n" % (
                "EA", "mt5", "py", "inrange", "exact", "tags", "first_div"))
            _rf.write("-" * 148 + "\n")
            def _tag_order(_t):
                return (1 if _t.split("(")[0] in ("PY_AFTER_END", "CLEAN", "UNCLASSIFIED") else 0, _t)
            for _rr in _rpt_rows:
                _ordered = sorted(_rr["tags"], key=_tag_order)
                _rf.write("%-52s %4d %4d %7d %5d  %-52s  %s\n" % (
                    _rr["ea"][:52], _rr["mt5"], _rr["py"], _rr["inrange"], _rr["exact"],
                    ",".join(_ordered)[:52], _rr["fdiv"][:80]))
            _rf.write("\n")

            # Section 3 — INDEX: EA name → per-rule folder (pull full detail without inlining)
            _rf.write("=== INDEX (EA -> folder) ===\n")
            for _rr in _rpt_rows:
                _rf.write("%-52s  %s\n" % (_rr["ea"][:52], _rr["sub"]))

        _d("PARITY REPORT: %s" % _report_path)

        # ── Eval Windows Report ──────────────────────────────────────────
        # WHY (June 2026): After parity comparison, auto-run sliding-window
        #      eval simulation for each rule. Answers "would this rule pass
        #      the Get Leveraged challenge starting on any given date?"
        # CHANGED: June 2026 — eval_windows_report feature
        try:
            from project3_live_trading.batch_compare_reports import generate_eval_windows_report
            _rules_dir = _py_rules_dir()
            if _rules_dir and os.path.isdir(_rules_dir):
                # CHANGED 2026-06-25 — combined report computes from MT5 trades in
                #   BOTH modes. Backtest mode keys mt5_by_source by the rule JSON
                #   filename (python.source) so the report can also load rd for
                #   labels. Scenario mode has no Python run (python.source is null)
                #   for every EA, so key by EA name instead — otherwise the map is
                #   empty and the report generates nothing.
                _has_py_src = any((ea.get("python") or {}).get("source")
                                  for ea in _consolidated.get("eas", []))
                _mt5_by_source = {}
                for ea in _consolidated.get("eas", []):
                    _nm = ea.get("name", "")
                    _src = (ea.get("python") or {}).get("source")
                    if _has_py_src:
                        if not _src:
                            continue
                        _key = _src      # backtest: rule JSON filename
                    else:
                        if not _nm:
                            continue
                        _key = _nm       # scenario: EA folder name
                    _csv_path = os.path.join(dump, _nm, "mt5_trades.csv")
                    # First EA wins per key (composites share a rule JSON).
                    _mt5_by_source.setdefault(_key, _csv_path)

                _eval_path = generate_eval_windows_report(
                    rules_dir=_rules_dir,
                    reports_dir=reports,
                    out_dir=dump,
                    only_sources=list(_mt5_by_source.keys()),
                    mt5_by_source=_mt5_by_source,
                    trade_source='mt5',
                    manifest_path=(mpath if os.path.isfile(mpath) else None),
                    firm_id='leveraged',
                    challenge_id='leveraged_standard',
                    account_size=(int(_acct_var.get())
                                  if (_acct_var is not None
                                      and str(_acct_var.get()).strip())
                                  else 10000),
                    risk_pct=1.0,
                    sl_pips=150.0,
                    pip_value_per_lot=1.0,
                )
                if _eval_path:
                    _d("EVAL WINDOWS REPORT: %s" % _eval_path)
                else:
                    _d("EVAL WINDOWS: no report generated (check logs)")
            else:
                _d("EVAL WINDOWS: rules_dir not found — skipping")
        except Exception as _eval_err:
            _d("EVAL WINDOWS ERROR: %r" % _eval_err)
    except Exception as _rpe:
        _d("PARITY_REPORT write failed: %r" % _rpe)

    if _dbg is not None:
        try: _dbg.close()
        except Exception: pass

    _append("DEBUG DUMP: %s" % dump)
    _append("DEBUG LOG: %s" % _dbg_path)
    try:
        os.startfile(dump)
    except Exception:
        pass


def _do_run_all(source_var):
    """Chain: Generate → Compile → Build → Run → Compare → Debug Dump."""
    def chain():
        _append("\n===== RUN ALL =====")
        if not _step_generate(source_var):
            _append("===== STOPPED (generate) =====")
            return
        if not _step_compile():
            _append("===== STOPPED (compile) =====")
            return
        if not _step_build_run_files():
            _append("===== STOPPED (build) =====")
            return
        if not _step_run_tests():
            _append("===== STOPPED (run) =====")
            return
        if not _step_compare():
            _append("===== STOPPED (compare) =====")
            return
        _write_debug_dump()
        _append("===== RUN ALL COMPLETE =====")
    _run_bg(chain)


# CHANGED: June 2026 — headless compile via metaeditor64; remembers the .exe path per machine
# WHY: user runs on two PCs — the path differs; save it once per hostname in gitignored
#   p1_config.json so neither machine inherits the other's path.
def _do_compile():
    _run_bg(_step_compile)


def _do_build_run_files():
    def work():
        try:
            from project3_live_trading.batch_ea_tools import emit_tester_inis
            # CHANGED: June 2026 — always recompute canonical (self-correcting); no dialogs.
            global _last_out_dir, _last_bat_path, _last_reports_dir
            batch_dir = _canonical_batch_dir()
            if not batch_dir:
                _append("Build run files cancelled (no MT5 data folder known).")
                return
            _last_out_dir = batch_dir
            manifest = os.path.join(batch_dir, "batch_manifest.json")
            if not os.path.isfile(manifest):
                _append("No manifest at %s — run '1. Generate EAs' first." % manifest)
                return
            data_dir = _derive_data_dir(batch_dir)
            if not data_dir:
                _append("Couldn't derive MT5 data dir from %s — pick it." % batch_dir)
                data_dir = filedialog.askdirectory(
                    title="Pick MT5 DATA folder (contains MQL5\\Experts)")
                if not data_dir:
                    _append("Build run files cancelled (no data dir).")
                    return
            # CHANGED: June 2026 — reports in a subfolder of Experts\batch, not the EA folder
            reports_dir = os.path.join(batch_dir, "reports")
            _append("Using batch dir : %s" % batch_dir)
            _append("Reports will be written to: %s" % os.path.abspath(reports_dir))
            # CHANGED: June 2026 — resolve terminal64 (derive from metaeditor → ask once → save)
            from project3_live_trading.batch_ea_tools import (
                resolve_terminal_path, save_terminal_path)
            term = resolve_terminal_path()
            if not term:
                _append("terminal64.exe not found automatically on this machine.")
                term = filedialog.askopenfilename(
                    title="Locate terminal64.exe — saved for THIS computer only",
                    filetypes=[("terminal64", "terminal64.exe"), ("All exe", "*.exe")])
                if term and save_terminal_path(term):
                    _append("Saved terminal64 path for this machine (%s)."
                            % __import__('socket').gethostname())
            emit_tester_inis(manifest, data_dir, "Experts\\batch", reports_dir,
                             terminal_exe=term or None)
            # CHANGED: June 2026 — record bat + reports dir for the Run Tests / Compare steps
            _last_bat_path    = os.path.join(batch_dir, "run_all_tests.bat")
            _last_reports_dir = reports_dir
            _append("Run files written.")
            _append("If you used '1b. Compile EAs', the .ex5 are already in MQL5\\Experts\\batch.")
            if term:
                _append("Terminal: %s" % term)
            _append("Next: click '2b. Run Tests' (make sure MT5 is closed).")
        except Exception as e:
            _append("ERROR during build run files: %s" % e)
    _run_bg(work)


def _do_run_tests():
    def work():
        try:
            import subprocess
            global _last_bat_path, _last_reports_dir
            bat = _last_bat_path
            if not bat or not os.path.isfile(bat):
                bat = filedialog.askopenfilename(
                    title="Pick run_all_tests.bat",
                    initialdir=(_last_out_dir or None),
                    filetypes=[("Batch", "*.bat")])
                if not bat:
                    _append("Run Tests cancelled (no .bat).")
                    return

            if _mt5_is_running():
                _append("MT5 is OPEN — close the MetaTrader 5 terminal first, then click "
                        "2b. Run Tests again. The Strategy Tester can't run while the "
                        "terminal holds the files.")
                return

            import subprocess, threading, time, glob as _glob

            # CHANGED: June 2026 — read total EA count from manifest for progress readout
            # WHY: show done/total and remaining so you know how many rules are left.
            _total_eas = 0
            try:
                import json as _json
                _man = os.path.join(os.path.dirname(bat), "batch_manifest.json")
                if os.path.isfile(_man):
                    with open(_man, encoding="utf-8") as _f:
                        _total_eas = len(_json.load(_f))
            except Exception:
                _total_eas = 0

            # CHANGED: June 2026 — clear old reports so the done/total count is accurate
            # WHY: stale reports from prior runs pollute the counter (5/4 done) and Compare.
            if _last_reports_dir and os.path.isdir(_last_reports_dir):
                _cleared = 0
                # CHANGED: June 2026 — clear all report formats including old .xlsx.htm double-ext files
                for _f in (_glob.glob(os.path.join(_last_reports_dir, "*.xlsx.htm")) +
                           _glob.glob(os.path.join(_last_reports_dir, "*.xlsx")) +
                           _glob.glob(os.path.join(_last_reports_dir, "*.htm")) +
                           _glob.glob(os.path.join(_last_reports_dir, "*.html"))):
                    try:
                        os.remove(_f)
                        _cleared += 1
                    except Exception:
                        pass
                if _cleared:
                    _append("Cleared %d old report(s) from %s" % (_cleared, _last_reports_dir))

            _append("Running backtests via %s ..." % os.path.basename(bat))
            if _total_eas:
                _append("Running %d EA(s). Each takes ~1-2 min headless." % _total_eas)
            _append("MT5 runs each EA headless (no visible window). Watch the heartbeat below.")

            proc = subprocess.Popen(
                ["cmd", "/c", bat],
                cwd=os.path.dirname(bat),
                stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                text=True, bufsize=1)

            # CHANGED: June 2026 — number the RUNNING lines as they stream so you know position
            # WHY: each RUNNING line gets tagged [k/N] so you see which EA is in progress.
            _seen = {"n": 0}
            def _pump():
                for line in proc.stdout:
                    line = line.rstrip()
                    if not line:
                        continue
                    if "=== RUNNING" in line:
                        _seen["n"] += 1
                        if _total_eas:
                            line = line.replace("=== RUNNING",
                                                "=== RUNNING [%d/%d]" % (_seen["n"], _total_eas))
                    _append("  " + line)
            _t = threading.Thread(target=_pump, daemon=True)
            _t.start()

            # Heartbeat: every 5s prove work is happening — MT5 alive + newest report growing.
            # CHANGED: June 2026 — heartbeat replaces silent wait; gives live proof per-EA
            rdir = _last_reports_dir
            _beat = 0
            while proc.poll() is None:
                time.sleep(5)
                _beat += 1
                alive = _mt5_is_running()
                newest = ""
                size = 0
                if rdir and os.path.isdir(rdir):
                    # CHANGED: June 2026 — tester now writes .xlsx (full deal data)
                    _files = (_glob.glob(os.path.join(rdir, "*.xlsx")) +
                              _glob.glob(os.path.join(rdir, "*.htm")) +
                              _glob.glob(os.path.join(rdir, "*.html")) +
                              _glob.glob(os.path.join(rdir, "*.xml")))
                    if _files:
                        newest = max(_files, key=os.path.getmtime)
                        try:
                            size = os.path.getsize(newest)
                        except OSError:
                            size = 0
                done_n = (len(_glob.glob(os.path.join(rdir, "*.xlsx"))) +
                          len(_glob.glob(os.path.join(rdir, "*.htm"))) +
                          len(_glob.glob(os.path.join(rdir, "*.html")))) if (rdir and os.path.isdir(rdir)) else 0
                # CHANGED: June 2026 — show done/total and remaining so you know how many left
                # WHY: clear progress readout (3/6 done, 3 left) instead of just count.
                if _total_eas:
                    _remaining = max(0, _total_eas - done_n)
                    _prog = "%d/%d done, %d left" % (done_n, _total_eas, _remaining)
                else:
                    _prog = "%d done" % done_n
                _append("  [heartbeat %d] MT5 %s | %s%s"
                        % (_beat,
                           "running" if alive else "starting/closing",
                           _prog,
                           (" | writing %s (%d KB)" % (os.path.basename(newest), size // 1024))
                           if newest else ""))
                # A2: tail the MT5 tester log for the "bars processed" line
                _tlog_dir = os.path.join(_derive_data_dir(_last_out_dir) or "", "Tester", "logs")
                if os.path.isdir(_tlog_dir):
                    _tlogs = _glob.glob(os.path.join(_tlog_dir, "*.log"))
                    if _tlogs:
                        _tl = max(_tlogs, key=os.path.getmtime)
                        try:
                            with open(_tl, "rb") as _f:
                                _f.seek(max(0, os.path.getsize(_tl) - 400))
                                _tail = _f.read().decode("utf-16", "ignore").strip().splitlines()
                            if _tail:
                                _append("    tester: " + _tail[-1][-120:])
                        except Exception:
                            pass
            _t.join(timeout=2)
            rc = proc.returncode

            if rc == 0:
                _append("All tester passes finished.")
            else:
                _append("Tester batch exited with code %d (some passes may have failed)." % rc)

            if _last_reports_dir and os.path.isdir(_last_reports_dir):
                # CHANGED: June 2026 — count .xlsx/.htm/.html (tester now writes .xlsx)
                n_rep = len([f for f in os.listdir(_last_reports_dir)
                             if f.lower().endswith((".xlsx", ".htm", ".html"))])
                # CHANGED: June 2026 — explicit done/total count in final summary
                _append("Done: %d/%d reports written." % (n_rep, _total_eas or n_rep))
                _append("Reports folder: %s" % _last_reports_dir)
                _append("Next: click '3. Compare Reports'.")
            else:
                _append("Done. Next: click '3. Compare Reports' and pick the reports folder.")
        except Exception as e:
            _append("ERROR during run tests: %s" % e)
    _run_bg(work)


def _do_compare():
    def work():
        try:
            from project3_live_trading.batch_compare_reports import compare_reports
            import glob as _g
            # CHANGED: June 2026 — auto-resolve ALL paths (reports, manifest); NO dialogs
            # CHANGED: June 2026 — Python trades now loaded from stored backtest files; no CSV dir
            # WHY: trades exist in backtest_trades_{TF}.json from backtesting; fully automatic.
            global _last_reports_dir, _last_out_dir

            # 1) REPORTS DIR — remembered from Run Tests/Build Run Files; else derive from batch dir
            reports = (_last_reports_dir
                       if (_last_reports_dir and os.path.isdir(_last_reports_dir)) else None)
            if not reports:
                bdir = _last_out_dir or _canonical_batch_dir()
                if bdir:
                    cand = os.path.join(bdir, "reports")
                    if os.path.isdir(cand):
                        reports = cand
            if not reports or not os.path.isdir(reports):
                _append("No reports folder found — run 2b. Run Tests first.")
                return

            # 2) MANIFEST — sits next to the batch dir; auto-resolve, never ask
            manifest = ""
            bdir = _last_out_dir or _canonical_batch_dir()
            if bdir:
                m = os.path.join(bdir, "batch_manifest.json")
                if os.path.isfile(m):
                    manifest = m
            # also try one level up from reports (…\batch\reports → …\batch)
            if not manifest:
                m2 = os.path.join(os.path.dirname(reports), "batch_manifest.json")
                if os.path.isfile(m2):
                    manifest = m2
            # manifest is optional for compare_reports; empty string is fine if not found

            # 3) PYTHON TRADES — pulled automatically from stored outputs/rules/*.json
            # WHY: trades already exist on disk from backtesting; no manual export needed.
            #   Each rule_*_{combo}_{exit}_{hash}_{TF}.json has a trades list.
            _append("Comparing reports in %s ..." % reports)
            if manifest:
                _append("Manifest: %s" % manifest)
            else:
                _append("(No manifest found — will use EA names only for matching)")
            _append("Python trades loaded from outputs/rules/*.json (automatic).")

            # No python_dir needed; trades come from outputs/rules/*.json via manifest
            rows = compare_reports(reports, "", manifest)
            for line in rows:
                _append(line)
            _append("comparison_summary.csv written in the reports folder.")
        except Exception as e:
            _append("ERROR during compare: %s" % e)
    _run_bg(work)


# ── panel builder ─────────────────────────────────────────────────────────────

def build_panel(parent):
    global _log, _grid_tree, _cur_source
    global _f_stage, _f_tf, _f_dir, _f_mintr, _f_minwr, _f_sort, _f_profitable, _nostats_lbl
    global _f_exit_vars, _f_exit_all, _f_exit_menu, _f_exit_menuobj
    panel = tk.Frame(parent, bg=BG)

    hdr = tk.Frame(panel, bg=DARK)
    hdr.pack(fill="x")
    tk.Label(hdr, text="Batch EAs", bg=DARK, fg="white",
             font=("Segoe UI", 13, "bold"), padx=12, pady=8).pack(side=tk.LEFT)

    # Source + action buttons row
    # WHY: default 'backtest' shows the last matrix run without mixing in saved rules.
    #   Use 'this run' after a backtest to scope to the run just launched; use 'all' to merge.
    # CHANGED: July 2026 — default 'backtest' (was 'all') + added 'this run' option
    src_var = tk.StringVar(value="backtest")
    bar = tk.Frame(panel, bg=BG)
    bar.pack(fill="x", padx=10, pady=(8, 2))

    # ── Mode toggle: Run Backtest (current behavior) vs Run Scenario (load the
    #    SAME discovery rules the Run Backtest panel loads, straight to MT5). ──
    # CHANGED: June 2026 — Run Backtest / Run Scenario mode toggle
    global _mode_var, _scenario_var, _scenario_paths, _scenario_combo, _src_combo_ref
    _mode_var = tk.StringVar(value="Run Backtest")
    tk.Label(bar, text="Mode:", bg=BG, fg=DARK,
             font=("Segoe UI", 9)).pack(side=tk.LEFT)
    mode_combo = ttk.Combobox(bar, textvariable=_mode_var, width=14, state="readonly",
                              values=["Run Backtest", "Run Scenario"])
    mode_combo.pack(side=tk.LEFT, padx=(4, 12))

    # Existing Source dropdown — shown only in Run Backtest mode.
    _src_lbl = tk.Label(bar, text="Source:", bg=BG, fg=DARK, font=("Segoe UI", 9))
    _src_lbl.pack(side=tk.LEFT)
    # CHANGED: June 2026 — add backtest + optimizer + all options
    src_combo = ttk.Combobox(bar, textvariable=src_var, width=14, state="readonly",
                              values=["backtest", "this run", "all", "optimizer", "saved_rules", "my_rules"])
    src_combo.pack(side=tk.LEFT, padx=6)
    src_combo.bind("<<ComboboxSelected>>", lambda e: _populate_grid(src_var.get()))
    _src_combo_ref = src_combo
    # WHY: src_var is local to build_panel; filter-widget lambdas use _cur_source (module
    #   global updated at the top of _populate_grid) so they always pass the active source.

    # New Scenario dropdown — shown only in Run Scenario mode (packed by _apply_mode).
    _scenario_var = tk.StringVar(value="")
    _scn_lbl = tk.Label(bar, text="Scenario:", bg=BG, fg=DARK, font=("Segoe UI", 9))
    _scenario_combo = ttk.Combobox(bar, textvariable=_scenario_var, width=42,
                                   state="readonly", values=["⏳ Loading…"])

    def _load_scenario_sources_async():
        # Scanner reads several JSON files; run off the UI thread, then push the
        # results + first-source grid back via .after(0, ...).
        from shared.scenario_sources import get_available_sources
        try:
            avail = get_available_sources()
        except Exception:
            avail = []
        labels = [s[0] for s in avail]
        _scenario_paths.clear()
        _scenario_paths.update({s[0]: s[1] for s in avail})

        def _ui():
            _scenario_combo['values'] = labels or ["(no sources found)"]
            if labels and not _scenario_var.get():
                _scenario_combo.set(labels[0])
            _populate_grid_scenario()
        try:
            _scenario_combo.after(0, _ui)
        except Exception:
            pass

    def _apply_mode(*_a):
        if _mode_var.get() == "Run Scenario":
            _src_lbl.pack_forget(); _src_combo_ref.pack_forget()
            _scn_lbl.pack(side=tk.LEFT, after=mode_combo)
            _scenario_combo.pack(side=tk.LEFT, padx=6, after=_scn_lbl)
            import threading as _t
            _t.Thread(target=_load_scenario_sources_async, daemon=True).start()
        else:
            _scn_lbl.pack_forget(); _scenario_combo.pack_forget()
            _src_lbl.pack(side=tk.LEFT, after=mode_combo)
            _src_combo_ref.pack(side=tk.LEFT, padx=6, after=_src_lbl)
            _populate_grid(src_var.get())

    mode_combo.bind("<<ComboboxSelected>>", _apply_mode)
    _scenario_combo.bind("<<ComboboxSelected>>", lambda e: _populate_grid_scenario())

    # Account-size dropdown — the fallback when a rule carries no account_size.
    # Applies in BOTH modes. Populated from the firm's account_sizes (5 tiers).
    global _acct_var
    _acct_var = tk.StringVar(value="10000")
    tk.Label(bar, text="Account:", bg=BG, fg=DARK,
             font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=(12, 2))

    def _acct_sizes():
        try:
            import json
            _p = os.path.join(
                os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
                "prop_firms", "leveraged.json")
            _d = json.load(open(_p, encoding="utf-8"))
            _szs = (_d.get("challenges", [{}]) or [{}])[0].get("account_sizes")
            if _szs:
                return [str(int(s)) for s in _szs]
        except Exception:
            pass
        return ["5000", "10000", "25000", "50000", "100000"]

    _acct_combo = ttk.Combobox(bar, textvariable=_acct_var, width=10,
                               state="readonly", values=_acct_sizes())
    _acct_combo.pack(side=tk.LEFT, padx=(0, 12))

    # CHANGED 2026-06-26 — action buttons get their own row. The Mode/Scenario/
    # Account dropdowns were consuming the top bar's width and pushing "▶ Run All"
    # off the right edge. Separate row = all 6 buttons always visible.
    btnbar = tk.Frame(panel, bg=BG)
    btnbar.pack(fill="x", padx=10, pady=(2, 2))

    def _btn(text, cmd):
        return tk.Button(btnbar, text=text, command=cmd, bg=MIDGREY, fg="white",
                         relief=tk.FLAT, cursor="hand2", padx=12, pady=4,
                         font=("Segoe UI", 9, "bold"))

    _btn("1. Generate EAs",    lambda: _do_generate(src_var)).pack(side=tk.LEFT, padx=4)
    # CHANGED: June 2026 — 1b runs metaeditor64 headlessly; no manual MetaEditor needed
    _btn("1b. Compile EAs",   _do_compile).pack(side=tk.LEFT, padx=4)
    _btn("2. Build Run Files", _do_build_run_files).pack(side=tk.LEFT, padx=4)
    # CHANGED: June 2026 — 2b launches run_all_tests.bat in-app; streams progress to log
    _btn("2b. Run Tests",      _do_run_tests).pack(side=tk.LEFT, padx=4)
    _btn("3. Compare Reports", _do_compare).pack(side=tk.LEFT, padx=4)
    # CHANGED: June 2026 — Run All chains all steps + writes debug dump
    _btn("▶ Run All", lambda: _do_run_all(src_var)).pack(side=tk.LEFT, padx=8)

    # CHANGED: June 2026 — filter row (parity with Strategy Refiner)
    filt_row = tk.Frame(panel, bg=BG)
    filt_row.pack(fill="x", padx=10, pady=(0, 4))

    # WHY: After saving new rules, the grid doesn't auto-update. This button
    #      reloads the full strategy list from disk so newly saved rules appear.
    # CHANGED: June 2026 — refresh button for batch grid
    tk.Button(filt_row, text="🔄 Refresh", font=("Segoe UI", 9, "bold"),
              bg=MIDGREY, fg="white", relief="raised", cursor="hand2",
              padx=6, pady=1,
              command=lambda: _populate_grid(_cur_source)).pack(side=tk.LEFT, padx=(0, 10))

    tk.Label(filt_row, text="Stage:", bg=BG, fg=DARK,
             font=("Segoe UI", 9)).pack(side=tk.LEFT)
    _f_stage = ttk.Combobox(filt_row, width=10, state="readonly",
                             values=["All", "Evaluation", "Funded", "Phase 1", "Phase 2"])
    _f_stage.set("All")
    _f_stage.pack(side=tk.LEFT, padx=(2, 6))

    tk.Label(filt_row, text="TF:", bg=BG, fg=DARK,
             font=("Segoe UI", 9)).pack(side=tk.LEFT)
    _f_tf = ttk.Combobox(filt_row, width=6, state="readonly",
                          values=["All", "M5", "M15", "H1", "H4", "D1"])
    _f_tf.set("All")
    _f_tf.pack(side=tk.LEFT, padx=(2, 6))

    tk.Label(filt_row, text="Dir:", bg=BG, fg=DARK,
             font=("Segoe UI", 9)).pack(side=tk.LEFT)
    _f_dir = ttk.Combobox(filt_row, width=6, state="readonly",
                           values=["All", "BUY", "SELL"])
    _f_dir.set("All")
    _f_dir.pack(side=tk.LEFT, padx=(2, 6))

    # CHANGED: June 2026 — multi-select exit-strategy filter via Menubutton
    # WHY: dropdown of checkboxes (All / specific exits) — compact, fits in filter row.
    # HOTFIX: June 2026 — store Menu object in _f_exit_menuobj (not _f_exit_menu['menu'], which is a str)
    tk.Label(filt_row, text="Exits:", bg=BG, fg=DARK,
             font=("Segoe UI", 9)).pack(side=tk.LEFT)
    _f_exit_menu = tk.Menubutton(filt_row, text="All ▾", relief=tk.RAISED, bg=WHITE,
                                 font=("Segoe UI", 9), padx=6, pady=2, cursor="hand2")
    _f_exit_menuobj = tk.Menu(_f_exit_menu, tearoff=0)  # keep the OBJECT
    _f_exit_menu.config(menu=_f_exit_menuobj)           # attach it
    _f_exit_menu.pack(side=tk.LEFT, padx=(2, 6))

    tk.Label(filt_row, text="Min Trades:", bg=BG, fg=DARK,
             font=("Segoe UI", 9)).pack(side=tk.LEFT)
    _f_mintr = tk.Entry(filt_row, width=5, font=("Segoe UI", 9))
    _f_mintr.pack(side=tk.LEFT, padx=(2, 6))

    tk.Label(filt_row, text="Min Win%:", bg=BG, fg=DARK,
             font=("Segoe UI", 9)).pack(side=tk.LEFT)
    _f_minwr = tk.Entry(filt_row, width=5, font=("Segoe UI", 9))
    _f_minwr.pack(side=tk.LEFT, padx=(2, 6))

    tk.Label(filt_row, text="Sort:", bg=BG, fg=DARK,
             font=("Segoe UI", 9)).pack(side=tk.LEFT)
    _f_sort = ttk.Combobox(filt_row, width=12, state="readonly",
                            values=["Prop Score", "Net Pips", "Win Rate", "PF", "Trades"])
    _f_sort.set("Net Pips")
    _f_sort.pack(side=tk.LEFT, padx=(2, 8))

    for _w in (_f_stage, _f_tf, _f_dir, _f_sort):
        _w.bind("<<ComboboxSelected>>", lambda e: _populate_grid(_cur_source))
    _f_mintr.bind("<Return>", lambda e: _populate_grid(_cur_source))
    _f_minwr.bind("<Return>", lambda e: _populate_grid(_cur_source))

    # CHANGED: June 2026 — profitable-only checkbox (parity with refiner, default ON)
    _f_profitable = tk.BooleanVar(value=True)
    tk.Checkbutton(filt_row, text="Profitable only", variable=_f_profitable,
                   bg=BG, fg=DARK, font=("Segoe UI", 9),
                   command=lambda: _populate_grid(_cur_source)).pack(side=tk.LEFT, padx=(4, 8))

    _nostats_lbl = tk.Label(filt_row, text="", bg=BG, fg="#888",
                             font=("Segoe UI", 8))
    _nostats_lbl.pack(side=tk.LEFT, padx=4)

    # CHANGED: June 2026 — grid + log in a vertical PanedWindow so the log box is drag-resizable
    # WHY: fixed-height log can't be resized; PanedWindow sash lets you drag to grow/shrink.
    _split = tk.PanedWindow(panel, orient="vertical", sashwidth=6,
                            sashrelief="raised", bg=BG)
    _split.pack(fill="both", expand=True, padx=10, pady=(0, 10))

    # CHANGED: June 2026 — rule selection grid with refiner-parity columns
    _grid_frame = tk.Frame(_split, bg=WHITE)
    # CHANGED: June 2026 — grid becomes first pane of PanedWindow (was pack with expand=True)

    # CHANGED: June 2026 — batch grid columns mirror the Rule Refiner strategy grid
    # CHANGED: June 2026 — added "profit" column (✅/❌/—) matching refiner's net>0 definition
    cols = ("sel", "id", "stage", "rule", "exit", "tf", "dir", "trades", "wr", "pf",
            "net_pips", "net_dollars", "net_pct", "avg_pips", "avg_dollars", "avg_pct",
            "profit", "win_pass", "prop_score", "off")
    _grid_tree = ttk.Treeview(_grid_frame, columns=cols, show="headings", height=10)
    _col_spec = [
        ("sel",        "☐",             34,  "center"),
        ("id",         "ID",            60,  "center"),
        ("stage",      "Stage",         70,  "center"),
        ("rule",       "Rule",         180,  "w"),
        ("exit",       "Exit Strategy",120,  "w"),
        ("tf",         "TF",            45,  "center"),
        ("dir",        "Dir",           50,  "center"),
        ("trades",     "Trades",        60,  "center"),
        ("wr",         "Win Rate",      70,  "center"),
        ("pf",         "PF",            55,  "center"),
        ("net_pips",   "Net Pips",      85,  "e"),
        ("net_dollars","Net $",         85,  "e"),
        ("net_pct",    "Net %",         65,  "e"),
        ("avg_pips",   "Avg Pips",      70,  "e"),
        ("avg_dollars","Avg $",         75,  "e"),
        ("avg_pct",    "Avg %",         65,  "e"),
        ("profit",     "Profit",        60,  "center"),
        ("win_pass",   "Win Pass",      95,  "center"),
        ("prop_score", "Prop Score",    80,  "center"),
        ("off",        "N/N+1",         60,  "center"),
    ]
    for c, t, w, a in _col_spec:
        _grid_tree.heading(c, text=t)
        # CHANGED: June 2026 — minwidth + stretch=False so columns hold their size and the
        #   horizontal scrollbar has real range (matches refiner grid feel)
        _grid_tree.column(c, width=w, minwidth=max(40, w // 2), anchor=a, stretch=False)
    _grid_tree.heading("sel", text="☐", command=_toggle_all)
    # CHANGED: June 2026 — row colors match refiner: green=profitable, red=losing, grey=no-data
    _grid_tree.tag_configure("profitable", foreground="#28a745")
    _grid_tree.tag_configure("losing",     foreground="#dc3545")
    _grid_tree.tag_configure("neutral",    foreground="#888888")
    _grid_tree.bind("<Button-1>", _toggle_row, add="+")
    _gsb = ttk.Scrollbar(_grid_frame, orient="vertical", command=_grid_tree.yview)
    _grid_tree.configure(yscrollcommand=_gsb.set)
    _gsb_h = ttk.Scrollbar(_grid_frame, orient="horizontal", command=_grid_tree.xview)
    _grid_tree.configure(xscrollcommand=_gsb_h.set)
    # CHANGED: June 2026 — pack scrollbars BEFORE the tree (refiner order) so each bar
    #   reserves its strip before the tree claims the remaining space.
    _gsb_h.pack(side="bottom", fill="x")   # h-bar first → owns the bottom strip
    _gsb.pack(side="right",  fill="y")     # v-bar next  → owns the right strip
    _grid_tree.pack(side="left", fill="both", expand=True)
    # CHANGED: June 2026 — wheel scroll on hover (parity with refiner); bind() not bind_all()
    #   so it only fires over this grid and doesn't hijack other panels' scrolling.
    def _grid_wheel(event):
        try:
            if event.num == 4:
                _grid_tree.yview_scroll(-3, "units")
            elif event.num == 5:
                _grid_tree.yview_scroll(3, "units")
            else:
                _grid_tree.yview_scroll(int(-1 * (event.delta / 120)) * 3, "units")
        except Exception:
            pass
        return "break"
    _grid_tree.bind("<MouseWheel>", _grid_wheel)
    _grid_tree.bind("<Button-4>",   _grid_wheel)
    _grid_tree.bind("<Button-5>",   _grid_wheel)
    _grid_tree.bind("<Shift-MouseWheel>",
                    lambda e: (_grid_tree.xview_scroll(int(-1*(e.delta/120))*3, "units"), "break")[1])

    # CHANGED: June 2026 — add grid frame to PanedWindow (was .pack(fill="both", expand=True))
    # WHY: minsize=120 prevents grid from collapsing; stretch="always" shares space with log.
    _split.add(_grid_frame, minsize=120, stretch="always")

    _populate_grid(src_var.get())

    # CHANGED: June 2026 — log becomes second pane of PanedWindow (was pack with height=6, expand=False)
    # WHY: drag the sash up → log grows taller, grid shrinks; drag down → reverse.
    _log = tk.Text(_split, font=("Consolas", 9), bg=WHITE, fg=DARK, wrap="word", height=8)
    _split.add(_log, minsize=60, stretch="always")
    _log.insert("end",
                "Batch flow: 1. Tick rules -> Generate EAs -> 1b. Compile EAs (headless) -> "
                "2. Build Run Files -> run .bat in MT5 -> 3. Compare Reports.\n\n")
    _log.configure(state="disabled")

    return panel


def refresh():
    pass
